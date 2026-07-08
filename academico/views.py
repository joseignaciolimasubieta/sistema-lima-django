from django.shortcuts import render, redirect, get_object_or_404
from django.db.models import Q, F, Sum, Count, Case, When, Value, IntegerField
from django.db.models.deletion import ProtectedError
from django.http import HttpResponse
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from django.template.loader import get_template
from xhtml2pdf import pisa
from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from .models import Honorario
import re
import os
import zipfile
from io import BytesIO
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
import json
from django.contrib import messages
from decimal import Decimal
import openpyxl
from itertools import chain
import datetime
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.db.models.functions import TruncMonth, TruncYear
from datetime import date, timedelta
from .models import Participante, Docente, Curso, Inscripcion, MovimientoCaja, CuentaCaja, Cliente, ServicioConsultora, Honorario, Empleado, PagoSueldo, VentaServicio, Asistencia, DatosEmpresa, Prestamo, PagoPrestamo, ArqueoCaja, CitaConsultora, ArchivoDigital, AnticipoEmpleado  # Importamos las tablas de la base de datos
from .forms import ParticipanteForm, DocenteForm, CursoForm, InscripcionForm, MovimientoCajaForm, ClienteForm, ServicioConsultoraForm, HonorarioForm, EmpleadoForm, PagoSueldoForm  # Importamos los formularios para crear entidades

def es_administrador(user):
       return user.is_superuser

@login_required
def dashboard(request):
    hoy = date.today()
    rango_fechas = request.GET.get('rango_fechas', '')

    # ==============================================================
    # OPTIMIZACIÓN CRÍTICA: FALLBACK AL MES ACTUAL
    # Evita ejecutar Inscripcion.objects.all() de toda la historia del sistema
    # cuando el usuario entra por primera vez, reduciendo la carga en la red.
    # ==============================================================
    if not rango_fechas:
        rango_fechas = hoy.strftime('%Y-%m')

    inscripciones_query = Inscripcion.objects.all()
    inscripciones_prev_query = Inscripcion.objects.none()
    tendencia_valida = False

    # --- MOTOR DE FILTRADO Y MÁQUINA DEL TIEMPO ---
    if rango_fechas:
        tendencia_valida = True
        if ' a ' in rango_fechas:
            partes = rango_fechas.split(' a ')
            fecha_inicio = partes[0].strip()
            fecha_fin = partes[1].strip() if len(partes) > 1 and partes[1].strip() else fecha_inicio
            
            inscripciones_query = inscripciones_query.filter(fecha_inscripcion__range=[fecha_inicio, fecha_fin])
            
            try:
                f_ini_obj = datetime.datetime.strptime(fecha_inicio, '%Y-%m-%d').date()
                f_fin_obj = datetime.datetime.strptime(fecha_fin, '%Y-%m-%d').date()
                delta = f_fin_obj - f_ini_obj + timedelta(days=1)
                inscripciones_prev_query = Inscripcion.objects.filter(
                    fecha_inscripcion__range=[f_ini_obj - delta, f_fin_obj - delta]
                )
            except Exception:
                tendencia_valida = False
            
        elif len(rango_fechas) == 4 and rango_fechas.isdigit():
            inscripciones_query = inscripciones_query.filter(fecha_inscripcion__year=rango_fechas)
            inscripciones_prev_query = Inscripcion.objects.filter(fecha_inscripcion__year=int(rango_fechas)-1)
            
        elif len(rango_fechas) == 7 and '-' in rango_fechas:
            try:
                anio, mes = map(int, rango_fechas.split('-'))
                inscripciones_query = inscripciones_query.filter(fecha_inscripcion__year=anio, fecha_inscripcion__month=mes)
                
                prev_mes = 12 if mes == 1 else mes - 1
                prev_anio = anio - 1 if mes == 1 else anio
                inscripciones_prev_query = Inscripcion.objects.filter(fecha_inscripcion__year=prev_anio, fecha_inscripcion__month=prev_mes)
            except Exception:
                tendencia_valida = False
        else:
            inscripciones_query = inscripciones_query.filter(fecha_inscripcion=rango_fechas)
            tendencia_valida = False

    # 2. MÉTRICAS GLOBALES FUSIONADAS (Evaluación directa en BD via SQL)
    datos_globales = inscripciones_query.aggregate(
        total_ins=Count('id'),
        suma_ing=Sum('importe')
    )
    
    datos_hoy = Inscripcion.objects.filter(fecha_inscripcion=hoy).aggregate(
        ins_hoy=Count('id'),
        ing_hoy=Sum('importe')
    )

    total_inscripciones = datos_globales['total_ins'] or 0
    suma_ingresos = datos_globales['suma_ing'] or Decimal('0.00')
    inscripciones_hoy = datos_hoy['ins_hoy'] or 0
    ingresos_hoy = datos_hoy['ing_hoy'] or Decimal('0.00')

    ticket_promedio = suma_ingresos / total_inscripciones if total_inscripciones > 0 else Decimal('0.00')
    
    total_cursos_activos = Curso.objects.filter(fecha_finalizacion__gte=hoy).count()

    # 3. REPORTE POR CURSO Y GRÁFICOS (Agrupación nativa en Supabase)
    reporte_cursos_bd = inscripciones_query.values('curso__nombre', 'modalidad').annotate(
        num_inscritos=Count('id'),
        total_generado=Sum('importe')
    ).order_by('-total_generado')

    prev_totales = {}
    if tendencia_valida:
        datos_prev = inscripciones_prev_query.values('curso__nombre', 'modalidad').annotate(
            total=Sum('importe')
        )
        prev_totales = { f"{d['curso__nombre']}_{d['modalidad']}": (d['total'] or Decimal('0.00')) for d in datos_prev }

    reporte_cursos = []
    for item in reporte_cursos_bd:
        actual = item['total_generado'] or Decimal('0.00')
        llave_busqueda = f"{item['curso__nombre']}_{item['modalidad']}"
        anterior = prev_totales.get(llave_busqueda, Decimal('0.00'))
        
        variacion = 0
        if anterior > 0:
            variacion = ((actual - anterior) / anterior) * 100
        elif actual > 0 and anterior == 0 and tendencia_valida:
            variacion = 100
            
        if not tendencia_valida or variacion == 0:
            item['tendencia_clase'] = 'bg-gray-50 text-gray-400'
            item['tendencia_texto'] = '■ ESTABLE'
        elif variacion > 0:
            item['tendencia_clase'] = 'bg-emerald-50 text-emerald-500'
            item['tendencia_texto'] = f'▲ {abs(variacion):.0f}%'
        else:
            item['tendencia_clase'] = 'bg-rose-50 text-rose-500'
            item['tendencia_texto'] = f'▼ {abs(variacion):.0f}%'

        reporte_cursos.append(item)

    # 4. CÁLCULO OPTIMIZADO DE CUENTAS OPERATIVAS (Mapeado directo en RAM)
    cuentas_operativas = CuentaCaja.objects.exclude(nombre__icontains='AHORRO').order_by('codigo')
    
    totales_bd = MovimientoCaja.objects.filter(
        cuenta__in=cuentas_operativas
    ).values('cuenta_id').annotate(
        t_entradas=Sum('monto', filter=Q(tipo='ENTRADA')),
        t_salidas=Sum('monto', filter=Q(tipo='SALIDA'))
    )
    
    mapa_saldos = {
        t['cuenta_id']: {
            'in': t['t_entradas'] or Decimal('0.00'),
            'out': t['t_salidas'] or Decimal('0.00')
        } for t in totales_bd
    }

    lista_cuentas_operativas = []
    total_operativo = Decimal('0.00')

    for cuenta in cuentas_operativas:
        datos_cuenta = mapa_saldos.get(cuenta.id, {'in': Decimal('0.00'), 'out': Decimal('0.00')})
        saldo_cuenta = datos_cuenta['in'] - datos_cuenta['out']
        
        cuenta.saldo_calculado = saldo_cuenta
        lista_cuentas_operativas.append(cuenta)
        total_operativo += saldo_cuenta

    contexto = {
        'rango_fechas': rango_fechas,
        'total_cursos': total_cursos_activos,
        'suma_ingresos': suma_ingresos,
        'inscripciones_hoy': inscripciones_hoy,
        'ingresos_hoy': ingresos_hoy,
        'ticket_promedio': ticket_promedio,
        'reporte_cursos': reporte_cursos,
        'total_operativo': total_operativo,
        'lista_cuentas_operativas': lista_cuentas_operativas,
    }
    
    return render(request, 'dashboard.html', contexto)

@login_required
def participantes(request):
    # En lugar de traer todos, traemos los últimos 150 para que cargue al instante
    lista_participantes = Participante.objects.all().order_by('-id')[:150]
    return render(request, 'participantes.html', {'participantes': lista_participantes})

@login_required
def crear_participante(request):
    if request.method == 'POST':
        # Si el usuario envió datos, los validamos y guardamos
        form = ParticipanteForm(request.POST)
        if form.is_valid():
            form.save() # ¡Esto lo guarda directo en Supabase!
            return redirect('participantes') # Lo devolvemos a la tabla
    else:
        # Si recién entra a la página, le mostramos el formulario vacío
        form = ParticipanteForm()
    
    return render(request, 'crear_participante.html', {'form': form})

@login_required
@user_passes_test(es_administrador)
def crear_docente(request):
    if request.method == 'POST':
        form = DocenteForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('cursos')
    else:
        form = DocenteForm()
    return render(request, 'crear_docente.html', {'form': form})

@login_required
def cursos(request):
    # Traemos los cursos ordenados por los más recientes
    lista_cursos = Curso.objects.select_related('docente').all().order_by('-id')
    
    # Capturamos lo que el usuario escriba en el buscador de texto
    buscar = request.GET.get('buscar', '')
    
    # ¡SOLUCIÓN!: Ahora atrapamos 'mes' (que es el name correcto del input en tu HTML)
    mes_busqueda = request.GET.get('mes', '') 
    
    # 1. FILTRO DE TEXTO
    if buscar:
        lista_cursos = lista_cursos.filter(
            Q(nombre__icontains=buscar) |
            Q(docente__nombre__icontains=buscar)
        )
        
    # 2. FILTRO DE CALENDARIO (MESES)
    if mes_busqueda:
        try:
            # El input envía "2026-06"
            anio, mes = mes_busqueda.split('-')
            
            # Filtramos los cursos cuya fecha de inicio sea en ese año y mes
            lista_cursos = lista_cursos.filter(
                fecha_inicio__year=anio, 
                fecha_inicio__month=mes
            )
        except ValueError:
            pass

    contexto = {
        'cursos': lista_cursos,
        'buscar': buscar,
        'mes_buscar': mes_busqueda # Mantenemos el valor para el placeholder
    }
    return render(request, 'cursos.html', contexto)

@login_required
@user_passes_test(es_administrador)
def crear_curso(request):
    if request.method == 'POST':
        form = CursoForm(request.POST, request.FILES)
        if form.is_valid():
            # 1. Guardamos el curso en la base de datos
            curso = form.save()
            
            # 2. Extracción segura del texto de duración (TU LÓGICA ORIGINAL)
            texto_duracion = str(
                getattr(curso, 'duracion', 
                getattr(curso, 'horario', 
                getattr(curso, 'duracion_curso', '')))
            ).lower()
            
            horas_numericas = 0
            carga_calculada = "0 horas"
            
            # 3. Cálculo automático de la carga horaria
            try:
                buscar_sesiones = re.search(r'(\d+)\s*(?:sesion|clase)', texto_duracion)
                if buscar_sesiones:
                    horas_numericas = int(buscar_sesiones.group(1))
                    carga_calculada = f"{horas_numericas} horas"
                else:
                    buscar_semanas = re.search(r'(\d+)\s*semana', texto_duracion)
                    if buscar_semanas:
                        semanas = int(buscar_semanas.group(1))
                        horas_numericas = semanas * 4
                        carga_calculada = f"{horas_numericas} horas"
            except Exception:
                horas_numericas = 0
                carga_calculada = "0 horas"

            # 4. Cálculo de Honorario (Horas * Bs. 22)
            monto_automatizado = horas_numericas * 22

            Honorario.objects.create(
                curso=curso,
                carga=carga_calculada,
                monto_acordado=monto_automatizado,
                anticipo=0.00,
                estado='PENDIENTE'
            )
            
            # ¡NUEVO! Dispara la ventanita azul de éxito
            messages.success(request, 'El curso fue registrado y sus honorarios calculados con éxito.')
            return redirect('cursos')
        else:
            messages.error(request, 'Ocurrió un error al registrar el curso. Verifique los datos ingresados.')
    else:
        form = CursoForm()
        
    return render(request, 'crear_curso.html', {'form': form, 'editando': False})


@login_required
@user_passes_test(es_administrador)
def editar_curso(request, id):
    curso = get_object_or_404(Curso, id=id) 
    
    if request.method == 'POST':
        form = CursoForm(request.POST, request.FILES, instance=curso)
        if form.is_valid():
            form.save()
            # ¡NUEVO! Dispara la ventanita azul de actualización
            messages.success(request, 'La información del curso se actualizó correctamente.')
            return redirect('cursos')
        else:
            messages.error(request, 'Ocurrió un error al actualizar. Verifique los datos.')
    else:
        form = CursoForm(instance=curso)
    
    context = {
        'form': form,
        'editando': True,
        'curso': curso
    }
    return render(request, 'crear_curso.html', context)


@login_required
@user_passes_test(es_administrador)
def eliminar_curso(request, id):
    # Usamos get_object_or_404 + ProtectedError para atrapar bloqueos (ej. si ya tiene honorarios pagados)
    curso = get_object_or_404(Curso, id=id)
    try:
        nombre_temporal = curso.nombre
        curso.delete()
        # ¡NUEVO! Dispara la ventanita roja de eliminación exitosa
        messages.success(request, f'El curso "{nombre_temporal}" fue eliminado del sistema.')
    except ProtectedError:
        messages.error(request, 'ACCIÓN BLOQUEADA: No se puede eliminar este curso porque ya tiene inscripciones asociadas.')
        
    return redirect('cursos')
@login_required
def inscripciones(request):
    buscar = request.GET.get('buscar', '')
    rango_fechas = request.GET.get('rango_fechas', '')

    # 1. Inicializamos Queries (Sin descargar nada aún)
    lista = Inscripcion.objects.select_related('participante', 'curso').filter(saldo_pendiente=0)
    ventas_extra = VentaServicio.objects.select_related('participante').all()

    # 2. Búsqueda de Texto
    if buscar:
        lista = lista.filter(
            Q(participante__nombre_completo__icontains=buscar) |
            Q(participante__celular__icontains=buscar) |
            Q(curso__nombre__icontains=buscar)
        )
        ventas_extra = ventas_extra.filter(
            Q(participante__nombre_completo__icontains=buscar) |
            Q(participante__celular__icontains=buscar) |
            Q(tipo_servicio__icontains=buscar) |
            Q(detalle__icontains=buscar)
        )
    # 3. Filtro de Fechas
    elif rango_fechas:
        if ' a ' in rango_fechas:
            fecha_inicio, fecha_fin = rango_fechas.split(' a ')
            lista = lista.filter(fecha_inscripcion__range=[fecha_inicio, fecha_fin])
            ventas_extra = ventas_extra.filter(fecha_venta__range=[fecha_inicio, fecha_fin])
        elif len(rango_fechas) == 4 and rango_fechas.isdigit():
            lista = lista.filter(fecha_inscripcion__year=rango_fechas)
            ventas_extra = ventas_extra.filter(fecha_venta__year=rango_fechas)
        elif len(rango_fechas) == 7 and '-' in rango_fechas:
            anio, mes = rango_fechas.split('-')
            lista = lista.filter(fecha_inscripcion__year=anio, fecha_inscripcion__month=mes)
            ventas_extra = ventas_extra.filter(fecha_venta__year=anio, fecha_venta__month=mes)
        else:
            lista = lista.filter(fecha_inscripcion=rango_fechas)
            ventas_extra = ventas_extra.filter(fecha_venta=rango_fechas)
    # OPTIMIZACIÓN CLAVE: Si no hay filtro, solo mostramos el MES ACTUAL para no colapsar la memoria
    else:
        hoy = datetime.date.today()
        lista = lista.filter(fecha_inscripcion__year=hoy.year, fecha_inscripcion__month=hoy.month)
        ventas_extra = ventas_extra.filter(fecha_venta__year=hoy.year, fecha_venta__month=hoy.month)

    # Ordenamiento en RAM, pero ahora solo de unos pocos registros (Ultra Rápido)
    def normalizar_fecha(obj):
        fecha = obj.fecha_inscripcion if hasattr(obj, 'fecha_inscripcion') else obj.fecha_venta
        return fecha.date() if isinstance(fecha, datetime.datetime) else (fecha or datetime.date.min)

    lista_combinada = sorted(chain(lista, ventas_extra), key=normalizar_fecha, reverse=False)

    return render(request, 'inscripciones.html', {
        'lista_combinada': lista_combinada, 
        'buscar': buscar,
        'rango_fechas': rango_fechas,
    })


@login_required
def crear_inscripcion(request):
    if request.method == 'POST':
        # 1. Extraemos los datos del formulario HTML
        nombre = request.POST.get('nombre_completo')
        celular = request.POST.get('celular')
        
        curso_id = request.POST.get('curso')
        fecha = request.POST.get('fecha_inscripcion')
        
        # Convertimos a Decimal de forma segura
        importe = Decimal(str(request.POST.get('importe', '0.00') or '0.00'))
        saldo_p = Decimal(str(request.POST.get('saldo_pendiente', '0.00') or '0.00'))
        
        banco = request.POST.get('banco')
        forma_pago = request.POST.get('forma_pago')
        modalidad = request.POST.get('modalidad').upper()
        vendedor = request.POST.get('vendedor').upper()
        registrado_por = request.POST.get('registrado_por').upper()
        
        # 2. Inteligencia: Buscamos si el alumno ya existe
        participantes_existentes = Participante.objects.filter(nombre_completo=nombre.upper())
        
        if participantes_existentes.exists():
            participante = participantes_existentes.first()
            participante.celular = celular
            participante.save()
        else:
            participante = Participante.objects.create(
                nombre_completo=nombre.upper(),
                celular=celular
            )
        
        # 3. Buscamos el curso o módulo seleccionado
        curso_seleccionado = Curso.objects.get(id=curso_id)
        subcursos = curso_seleccionado.subcursos.all().order_by('id')
        
        # =========================================================
        # 🤖 CEREBRO DE DISTRIBUCIÓN MATEMÁTICA (COMO EN EL EXCEL)
        # =========================================================
        if subcursos.exists():
            # ES UN MÓDULO: Dividimos el dinero entre sus cursos internos
            cantidad = subcursos.count()
            
            # Calculamos la base para no perder centavos
            importe_base = round(importe / cantidad, 2)
            saldo_base = round(saldo_p / cantidad, 2)
            
            importe_acumulado = Decimal('0.00')
            saldo_acumulado = Decimal('0.00')
            
            for i, subcurso in enumerate(subcursos):
                # El último curso absorbe la diferencia exacta (Ej: el de 99 Bs.)
                if i == cantidad - 1:
                    importe_final = importe - importe_acumulado
                    saldo_final = saldo_p - saldo_acumulado
                else:
                    importe_final = importe_base
                    saldo_final = saldo_base
                    importe_acumulado += importe_base
                    saldo_acumulado += saldo_base
                    
                # Validamos duplicados individuales
                if not Inscripcion.objects.filter(participante=participante, curso=subcurso).exists():
                    Inscripcion.objects.create(
                        participante=participante,
                        curso=subcurso, # Guardamos el CURSO, ignoramos el Módulo
                        fecha_inscripcion=fecha,
                        importe=importe_final,
                        saldo_pendiente=saldo_final,
                        banco=banco,
                        forma_pago=forma_pago,
                        modalidad=modalidad,
                        vendedor=vendedor,
                        registrado_por=registrado_por
                    )
            messages.success(request, f'Alumno inscrito con éxito en los {cantidad} cursos correspondientes al {curso_seleccionado.nombre}.')
            
        else:
            # ES UN CURSO NORMAL (SIN HIJOS)
            inscripcion_duplicada = Inscripcion.objects.filter(
                participante=participante, curso=curso_seleccionado, modalidad=modalidad
            ).exists()
            
            if inscripcion_duplicada:
                messages.error(request, f'ACCIÓN BLOQUEADA: El alumno {participante.nombre_completo} ya está registrado en el curso "{curso_seleccionado.nombre}".')
                return redirect('inscripciones')
                
            Inscripcion.objects.create(
                participante=participante,
                curso=curso_seleccionado,
                fecha_inscripcion=fecha,
                importe=importe,
                saldo_pendiente=saldo_p,
                banco=banco,
                forma_pago=forma_pago,
                modalidad=modalidad,
                vendedor=vendedor,
                registrado_por=registrado_por
            )
            messages.success(request, 'Participante registrado con éxito.')

        return redirect('inscripciones')
    
    else:
        hoy = date.today()
        cursos = Curso.objects.select_related('docente').annotate(
            orden_estado=Case(
                When(fecha_inicio__gt=hoy, then=Value(1)),
                When(fecha_inicio__lte=hoy, fecha_finalizacion__gte=hoy, then=Value(2)),
                default=Value(3),
                output_field=IntegerField(),
            )
        ).order_by('orden_estado', '-fecha_inicio')
        
        return render(request, 'crear_inscripcion.html', {'cursos': cursos, 'hoy': hoy})

@login_required
@user_passes_test(es_administrador)
def flujo_caja(request):
    rango_fechas = request.GET.get('rango_fechas', '')
    
    # --- OPTIMIZACIÓN 1: select_related('cuenta') ---
    # Esto trae toda la información de las cuentas en UN SOLO viaje a la BD, 
    # evitando hacer un viaje nuevo por cada fila de la tabla en el HTML.
    movimientos_qs = MovimientoCaja.objects.select_related('cuenta').all().order_by('fecha', 'id')
    
    if rango_fechas:
        if ' a ' in rango_fechas:
            fecha_inicio, fecha_fin = rango_fechas.split(' a ')
            movimientos_qs = movimientos_qs.filter(fecha__range=[fecha_inicio, fecha_fin])
        elif len(rango_fechas) == 4 and rango_fechas.isdigit():
            movimientos_qs = movimientos_qs.filter(fecha__year=rango_fechas)
        elif len(rango_fechas) == 7 and '-' in rango_fechas:
            anio, mes = rango_fechas.split('-')
            movimientos_qs = movimientos_qs.filter(fecha__year=anio, fecha__month=mes)
        else:
            movimientos_qs = movimientos_qs.filter(fecha=rango_fechas)
            
    saldo_corriente = 0
    lista_movimientos = []
    
    for mov in movimientos_qs:
        if mov.tipo == 'ENTRADA':
            saldo_corriente += mov.monto
        elif mov.tipo == 'SALIDA':
            saldo_corriente -= mov.monto
            
        mov.saldo_calculado = saldo_corriente  
        lista_movimientos.append(mov)
        
    # --- OPTIMIZACIÓN 2: AGREGACIÓN MASIVA EN BASE DE DATOS ---
    # En lugar de hacer un bucle 'for' que consulte la base de datos 20 veces, 
    # hacemos UNA SOLA consulta que agrupa y suma el dinero de todas las cuentas.
    totales_bd = MovimientoCaja.objects.values('cuenta_id').annotate(
        t_entradas=Sum('monto', filter=Q(tipo='ENTRADA')),
        t_salidas=Sum('monto', filter=Q(tipo='SALIDA'))
    )
    
    # Guardamos los resultados en la memoria RAM (diccionario) para acceso inmediato
    mapa_saldos = {}
    for t in totales_bd:
        mapa_saldos[t['cuenta_id']] = {
            'in': t['t_entradas'] or Decimal('0.00'),
            'out': t['t_salidas'] or Decimal('0.00')
        }

    cuentas_caja = CuentaCaja.objects.all().order_by('codigo')
    
    cuentas_operativas = []
    cuentas_ahorro = []
    total_operativo = Decimal('0.00')
    total_ahorro = Decimal('0.00')
    
    # Ahora el bucle solo lee de la RAM, no de la base de datos = Carga instantánea
    for cuenta in cuentas_caja:
        datos_cuenta = mapa_saldos.get(cuenta.id, {'in': Decimal('0.00'), 'out': Decimal('0.00')})
        cuenta.saldo_actual = datos_cuenta['in'] - datos_cuenta['out']
        
        if cuenta.codigo in ['001', '002', '003', '004']:
            cuentas_operativas.append(cuenta)
            total_operativo += cuenta.saldo_actual
        elif cuenta.codigo in ['005', '006', '007']:
            cuentas_ahorro.append(cuenta)
            total_ahorro += cuenta.saldo_actual

    total_absoluto = total_operativo + total_ahorro
    
    # Obtenemos los totales generales sumando directamente de la RAM
    total_entradas = sum(item['in'] for item in mapa_saldos.values())
    total_salidas = sum(item['out'] for item in mapa_saldos.values())
    saldo_total = total_entradas - total_salidas

    contexto = {
        'movimientos': lista_movimientos,
        'rango_fechas': rango_fechas,
        'total_entradas': total_entradas,
        'total_salidas': total_salidas,
        'saldo_total': saldo_total,
        'cuentas_caja': cuentas_caja,
        'cuentas_operativas': cuentas_operativas,
        'cuentas_ahorro': cuentas_ahorro,
        'total_operativo': total_operativo,
        'total_ahorro': total_ahorro,
        'total_absoluto': total_absoluto,
    }

    return render(request, 'flujo_caja.html', contexto)
@login_required
@user_passes_test(es_administrador)
def crear_movimiento(request):
    if request.method == 'POST':
        form = MovimientoCajaForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('flujo_caja')
    else:
        form = MovimientoCajaForm()
    return render(request, 'crear_movimiento.html', {'form': form})

@login_required
def consultora(request):
    # Traemos todos los servicios
    servicios = ServicioConsultora.objects.select_related('cliente').all().order_by('-fecha')
    
    # Atrapamos los parámetros enviados desde el buscador HTML
    buscar_cliente = request.GET.get('buscar_cliente', '').strip()
    mes_buscar = request.GET.get('mes_buscar', '').strip()
    
    # 1. Filtro por Razón Social o NIT del cliente
    if buscar_cliente:
        from django.db.models import Q
        servicios = servicios.filter(
            Q(cliente__nombre_contribuyente__icontains=buscar_cliente) |
            Q(cliente__nit__icontains=buscar_cliente)
        )
        
    # 2. Filtro por Periodo de Calendario (Ej: "2026-06")
    if mes_buscar:
        try:
            anio, mes = mes_buscar.split('-')
            # Filtra coincidiendo el año y el mes exacto del trámite
            servicios = servicios.filter(fecha__year=anio, fecha__month=mes)
        except ValueError:
            pass
            
    contexto = {
        'servicios': servicios,
        'buscar_cliente': buscar_cliente,
        'mes_buscar': mes_buscar
    }
    
    return render(request, 'consultora.html', contexto)

@login_required
def eliminar_servicio(request, servicio_id):
    servicio = get_object_or_404(ServicioConsultora, id=servicio_id)
    try:
        servicio.delete()
        messages.success(request, 'El trámite/servicio ha sido eliminado correctamente.')
    except ProtectedError:
        messages.error(request, 'No se puede eliminar este registro porque tiene movimientos relacionados.')
    return redirect('consultora')

@login_required
@user_passes_test(es_administrador)
def crear_servicio(request):
    if request.method == 'POST':
        post_data = request.POST.copy()
        nombre_express = post_data.get('nombre_express')
        cliente_id = post_data.get('cliente')
        
        # 1. LÓGICA EXPRESS DE CLIENTES
        if nombre_express and nombre_express.strip() != '':
            nit_express = post_data.get('nit_express', '').strip()
            celular_express = post_data.get('celular_express', '').strip()
            
            try:
                nuevo_cliente = Cliente.objects.create(
                    nombre=nombre_express.strip().upper(),
                    nit=nit_express if nit_express else None,
                    celular=celular_express if celular_express else None
                )
            except TypeError:
                try:
                    nuevo_cliente = Cliente.objects.create(
                        razon_social=nombre_express.strip().upper(),
                        nit=nit_express if nit_express else None,
                        celular=celular_express if celular_express else None
                    )
                except TypeError:
                    nuevo_cliente = Cliente.objects.create(
                        nombre_contribuyente=nombre_express.strip().upper(),
                        nit=nit_express if nit_express else None,
                        celular=celular_express if celular_express else None
                    )
            
            cliente_id = nuevo_cliente.id
            post_data['cliente'] = str(cliente_id)

        if not cliente_id or str(cliente_id).strip() == '':
            messages.error(request, 'Debe seleccionar un cliente de la lista o escribir los datos de uno nuevo.')
            return redirect('crear_servicio')

        # --- NUEVO: INYECCIÓN DE COMISIÓN PARA PASAR LA VALIDACIÓN ---
        # Engañamos a Django poniendo un 0. El 20% real se calculará en models.py
        if 'comision' not in post_data or not post_data['comision']:
            post_data['comision'] = '0'

        # 3. BLINDAJE DE GUARDADO
        form = ServicioConsultoraForm(post_data, request.FILES)
        if form.is_valid():
            servicio = form.save(commit=False)
            servicio.cliente = Cliente.objects.get(id=int(cliente_id))
            servicio.save() # Aquí ocurre la magia del 20% de models.py
            
            messages.success(request, 'Servicio/Consulta registrado exitosamente.')
            return redirect('consultora')
        else:
            messages.error(request, 'Ocurrió un error al guardar. Verifique los datos ingresados.')
            
    else:
        form = ServicioConsultoraForm()
        
    return render(request, 'crear_servicio.html', {'form': form})
@login_required
@user_passes_test(es_administrador)
def honorarios(request):
    lista_honorarios = Honorario.objects.select_related('curso__docente').all().order_by('-id')
    hoy = date.today()
    
    # 1. CAPTURAMOS LOS DATOS DEL NUEVO FORMULARIO
    busqueda_docente = request.GET.get('docente', '')
    rango_fechas = request.GET.get('rango_fechas', '') # Recibimos el nuevo calendario

    # 2. APLICAMOS LOS FILTROS
    if busqueda_docente:
        lista_honorarios = lista_honorarios.filter(curso__docente__nombre__icontains=busqueda_docente)
    
    # Lógica del Calendario Avanzado (Rango, Año, Mes o Día)
    if rango_fechas:
        # 1. Si eligió un rango con el mouse (Ej: "2026-06-01 a 2026-06-30")
        if ' a ' in rango_fechas:
            fecha_inicio, fecha_fin = rango_fechas.split(' a ')
            lista_honorarios = lista_honorarios.filter(curso__fecha_inicio__range=[fecha_inicio, fecha_fin])
            
        # 2. Si escribió solo el Año con el teclado (Ej: "2026")
        elif len(rango_fechas) == 4 and rango_fechas.isdigit():
            lista_honorarios = lista_honorarios.filter(curso__fecha_inicio__year=rango_fechas)
            
        # 3. Si escribió el Año y el Mes con el teclado (Ej: "2026-06")
        elif len(rango_fechas) == 7 and '-' in rango_fechas:
            anio, mes = rango_fechas.split('-')
            lista_honorarios = lista_honorarios.filter(curso__fecha_inicio__year=anio, curso__fecha_inicio__month=mes)
            
        # 4. Si hizo clic en un solo día exacto
        else:
            lista_honorarios = lista_honorarios.filter(curso__fecha_inicio=rango_fechas)

    # 3. Calculamos los estados dinámicos
    for h in lista_honorarios:
        if h.curso and h.curso.fecha_inicio and h.curso.fecha_finalizacion:
            if hoy < h.curso.fecha_inicio:
                h.estado_curso_calc = "NO INICIÓ"
            elif hoy > h.curso.fecha_finalizacion:
                h.estado_curso_calc = "FINALIZADO"
            else:
                h.estado_curso_calc = "EN CURSO"
        else:
            h.estado_curso_calc = "Fechas pendientes"
            
    # 4. Sumas financieras dinámicas y reales
    # Suma de cursos 100% liquidados
    total_pagado = lista_honorarios.filter(estado='PAGADO').aggregate(Sum('monto_acordado'))['monto_acordado__sum'] or 0
    
    # Suma del dinero que ya salió como anticipo en cursos que aún no terminan
    total_anticipos = lista_honorarios.filter(estado='PENDIENTE').aggregate(Sum('anticipo'))['anticipo__sum'] or 0
    
    # DEUDA REAL: Calcula (Monto Acordado - Anticipo) solo de los pendientes
    total_pendiente = lista_honorarios.filter(estado='PENDIENTE').aggregate(
        deuda_real=Sum(F('monto_acordado') - F('anticipo'))
    )['deuda_real'] or 0
    
    contexto = {
        'honorarios': lista_honorarios,
        'total_pendiente': total_pendiente,
        'total_anticipos': total_anticipos, # ¡Pasamos la nueva variable al HTML!
        'total_pagado': total_pagado,
        'busqueda_docente': busqueda_docente,
        'rango_fechas': rango_fechas,
    }
    return render(request, 'honorarios.html', contexto)

@login_required
def imprimir_recibo(request, inscripcion_id):
    # Buscamos la inscripción exacta que queremos imprimir
    inscripcion = get_object_or_404(Inscripcion, id=inscripcion_id)
    
    # Cargamos el diseño HTML del recibo
    template = get_template('recibo_pdf.html')
    contexto = {'inscripcion': inscripcion}
    html = template.render(contexto)
    
    # Preparamos la respuesta como un archivo PDF
    response = HttpResponse(content_type='application/pdf')
    # "inline" permite verlo en el navegador. Si quisieras forzar la descarga, usarías "attachment;"
    response['Content-Disposition'] = f'inline; filename="Recibo_LIMA_{inscripcion.participante.nombre_completo}.pdf"'
    
    # Convertimos el HTML a PDF
    pisa_status = pisa.CreatePDF(html, dest=response)
    
    if pisa_status.err:
        return HttpResponse('Hubo un error al generar el recibo PDF.')
    
    return response

@login_required
@user_passes_test(es_administrador)
def planillas(request):
    empresa = DatosEmpresa.objects.first()
    pagos = PagoSueldo.objects.select_related('empleado').all().order_by('-fecha_pago')
    
    mes_buscar = request.GET.get('mes_buscar', '')
    buscar = request.GET.get('buscar', '').strip() # <-- NUEVO: Atrapa el texto a buscar
    mes_literal_buscar = "" 
    
    # --- 1. FILTRO DE TEXTO (Nombre o C.I.) ---
    if buscar:
        pagos = pagos.filter(
            Q(empleado__nombre_completo__icontains=buscar) |
            Q(empleado__ci__icontains=buscar)
        )

    # --- 2. FILTRO DE MES ---
    if mes_buscar:
        pagos = pagos.filter(mes_correspondiente=mes_buscar)
        
        # Traductor para el encabezado de la página cuando se filtra
        meses_dict = {
            '01': 'ENERO', '02': 'FEBRERO', '03': 'MARZO', '04': 'ABRIL', 
            '05': 'MAYO', '06': 'JUNIO', '07': 'JULIO', '08': 'AGOSTO', 
            '09': 'SEPTIEMBRE', '10': 'OCTUBRE', '11': 'NOVIEMBRE', '12': 'DICIEMBRE'
        }
        try:
            anio, mes_num = str(mes_buscar).strip().split('-')
            mes_literal_buscar = f"{meses_dict.get(mes_num, '')} {anio}"
        except:
            mes_literal_buscar = str(mes_buscar).upper()
        
    totales = pagos.aggregate(
        t_salario_base=Sum('salario_base'),
        t_bono_antiguedad=Sum('bono_antiguedad'),
        t_bono_ventas=Sum('bono_ventas'),
        t_comisiones_certificados=Sum('comisiones_certificados'),
        t_bono_consultora=Sum('bono_consultora'),
        t_otros_bonos=Sum('otros_bonos'),
        t_aportes_afp=Sum('aportes_afp'),
        t_rc_iva=Sum('rc_iva'),
        t_anticipos=Sum('anticipos'),
        t_prestamos=Sum('prestamos'),
        t_multas=Sum('multas'),
        t_rendicion_cuentas=Sum('rendicion_cuentas'),
        t_pasanaku=Sum('pasanaku')
    )
    
    t_salario_base = totales['t_salario_base'] or Decimal('0.00')
    t_bono_antiguedad = totales['t_bono_antiguedad'] or Decimal('0.00')
    t_bono_ventas = totales['t_bono_ventas'] or Decimal('0.00')
    t_comisiones_certificados = totales['t_comisiones_certificados'] or Decimal('0.00')
    t_bono_consultora = totales['t_bono_consultora'] or Decimal('0.00')
    t_otros_bonos = totales['t_otros_bonos'] or Decimal('0.00')
    t_aportes_afp = totales['t_aportes_afp'] or Decimal('0.00')
    t_rc_iva = totales['t_rc_iva'] or Decimal('0.00')
    t_anticipos = totales['t_anticipos'] or Decimal('0.00')
    t_prestamos = totales['t_prestamos'] or Decimal('0.00')
    t_multas = totales['t_multas'] or Decimal('0.00')
    t_rendicion_cuentas = totales['t_rendicion_cuentas'] or Decimal('0.00')
    t_pasanaku = totales['t_pasanaku'] or Decimal('0.00')
    
    t_otros_bonos_combinados = t_comisiones_certificados + t_bono_consultora + t_otros_bonos
    t_total_ganado = t_salario_base + t_bono_antiguedad + t_bono_ventas + t_otros_bonos_combinados
    
    t_otros_descuentos_combinados = t_anticipos + t_prestamos + t_multas + t_rendicion_cuentas + t_pasanaku
    t_total_descuentos = t_aportes_afp + t_rc_iva + t_otros_descuentos_combinados
    t_liquido_pagable = t_total_ganado - t_total_descuentos
    
    contexto = {
        'pagos': pagos,
        'empresa': empresa,
        'mes_buscar': mes_buscar,
        'buscar': buscar, # <-- Enviamos la variable de búsqueda al HTML
        'mes_literal_buscar': mes_literal_buscar, 
        't_salario_base': t_salario_base,
        't_bono_antiguedad': t_bono_antiguedad,
        't_bono_ventas': t_bono_ventas,
        't_otros_bonos_combinados': t_otros_bonos_combinados,
        't_total_ganado': t_total_ganado,
        't_aportes_afp': t_aportes_afp,
        't_rc_iva': t_rc_iva,
        't_otros_descuentos_combinados': t_otros_descuentos_combinados,
        't_total_descuentos': t_total_descuentos,
        't_liquido_pagable': t_liquido_pagable,
    }
    return render(request, 'planillas.html', contexto)

@login_required
@user_passes_test(es_administrador)
def descargar_pdf_planilla(request):
    empresa = DatosEmpresa.objects.first()
    pagos = PagoSueldo.objects.select_related('empleado').all().order_by('-fecha_pago')
    
    mes_buscar = request.GET.get('mes_buscar', '')
    buscar = request.GET.get('buscar', '').strip() # <-- NUEVO: Atrapamos el texto buscado
    
    # --- NUEVO: APLICAMOS EL FILTRO DE TEXTO AL PDF ---
    if buscar:
        pagos = pagos.filter(
            Q(empleado__nombre_completo__icontains=buscar) |
            Q(empleado__ci__icontains=buscar)
        )

    if mes_buscar:
        pagos = pagos.filter(mes_correspondiente=mes_buscar)
        
    # --- LÓGICA: Separar Mes en Literal y Año ---
    mes_nombre = ""
    anio_nombre = "2026"  # Año por defecto seguro
    
    periodo_referencia = mes_buscar
    if not periodo_referencia and pagos.exists():
        periodo_referencia = pagos.first().mes_correspondiente
        
    if periodo_referencia:
        meses_dict = {
            '01': 'ENERO', '02': 'FEBRERO', '03': 'MARZO', '04': 'ABRIL',
            '05': 'MAYO', '06': 'JUNIO', '07': 'JULIO', '08': 'AGOSTO',
            '09': 'SEPTIEMBRE', '10': 'OCTUBRE', '11': 'NOVIEMBRE', '12': 'DICIEMBRE'
        }
        try:
            anio_num, mes_num = str(periodo_referencia).strip().split('-')
            mes_nombre = meses_dict.get(mes_num, '').upper()
            anio_nombre = anio_num
        except:
            pass

    totales = pagos.aggregate(
        t_salario_base=Sum('salario_base'),
        t_bono_antiguedad=Sum('bono_antiguedad'),
        t_bono_ventas=Sum('bono_ventas'),
        t_comisiones_certificados=Sum('comisiones_certificados'),
        t_bono_consultora=Sum('bono_consultora'),
        t_otros_bonos=Sum('otros_bonos'),
        t_aportes_afp=Sum('aportes_afp'),
        t_rc_iva=Sum('rc_iva'),
        t_anticipos=Sum('anticipos'),
        t_prestamos=Sum('prestamos'),
        t_multas=Sum('multas'),
        t_rendicion_cuentas=Sum('rendicion_cuentas'),
        t_pasanaku=Sum('pasanaku')
    )
    
    t_salario_base = totales['t_salario_base'] or Decimal('0.00')
    t_bono_antiguedad = totales['t_bono_antiguedad'] or Decimal('0.00')
    t_bono_ventas = totales['t_bono_ventas'] or Decimal('0.00')
    t_comisiones_certificados = totales['t_comisiones_certificados'] or Decimal('0.00')
    t_bono_consultora = totales['t_bono_consultora'] or Decimal('0.00')
    t_otros_bonos = totales['t_otros_bonos'] or Decimal('0.00')
    t_aportes_afp = totales['t_aportes_afp'] or Decimal('0.00')
    t_rc_iva = totales['t_rc_iva'] or Decimal('0.00')
    t_anticipos = totales['t_anticipos'] or Decimal('0.00')
    t_prestamos = totales['t_prestamos'] or Decimal('0.00')
    t_multas = totales['t_multas'] or Decimal('0.00')
    t_rendicion_cuentas = totales['t_rendicion_cuentas'] or Decimal('0.00')
    t_pasanaku = totales['t_pasanaku'] or Decimal('0.00')
    
    t_otros_bonos_combinados = t_comisiones_certificados + t_bono_consultora + t_otros_bonos
    t_total_ganado = t_salario_base + t_bono_antiguedad + t_bono_ventas + t_otros_bonos_combinados
    
    t_otros_descuentos_combinados = t_anticipos + t_prestamos + t_multas + t_rendicion_cuentas + t_pasanaku
    t_total_descuentos = t_aportes_afp + t_rc_iva + t_otros_descuentos_combinados
    t_liquido_pagable = t_total_ganado - t_total_descuentos
    
    contexto = {
        'pagos': pagos,
        'empresa': empresa,
        'mes_nombre': mes_nombre,      # <-- NUEVO
        'anio_nombre': anio_nombre,    # <-- NUEVO
        't_salario_base': t_salario_base,
        't_bono_antiguedad': t_bono_antiguedad,
        't_bono_ventas': t_bono_ventas,
        't_otros_bonos_combinados': t_otros_bonos_combinados,
        't_total_ganado': t_total_ganado,
        't_aportes_afp': t_aportes_afp,
        't_rc_iva': t_rc_iva,
        't_otros_descuentos_combinados': t_otros_descuentos_combinados,
        't_total_descuentos': t_total_descuentos,
        't_liquido_pagable': t_liquido_pagable,
    }
    
    template = get_template('pdf_planilla.html')
    html = template.render(contexto)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'inline; filename="Planilla_Ministerio_Sueldos.pdf"'
    
    pisa_status = pisa.CreatePDF(html, dest=response)
    
    if pisa_status.err:
        return HttpResponse('Hubo un error al generar el PDF de la planilla.')
        
    return response
@login_required
@user_passes_test(es_administrador)
def crear_pago(request):
    if request.method == 'POST':
        from decimal import Decimal
        from django.urls import reverse 
        from datetime import date
        
        # Función blindada contra errores de texto o vacíos
        def a_decimal(valor):
            if not valor or str(valor).strip() == 'None' or str(valor).strip() == '':
                return Decimal('0.00')
            try:
                return Decimal(str(valor).strip())
            except:
                return Decimal('0.00')

        # Protección para la cuenta de origen
        cuenta_origen_id = request.POST.get('cuenta_origen')
        if not cuenta_origen_id or str(cuenta_origen_id).strip() == '':
            cuenta_origen_id = None

        # Protección para la fecha
        fecha_recibida = request.POST.get('fecha_pago')
        if not fecha_recibida or str(fecha_recibida).strip() == '':
            fecha_recibida = date.today().strftime('%Y-%m-%d')

        pago = PagoSueldo(
            empleado_id=request.POST.get('empleado_id'),
            fecha_pago=fecha_recibida,
            mes_correspondiente=request.POST.get('mes_correspondiente'),
            
            salario_base=a_decimal(request.POST.get('salario_base')),
            bono_antiguedad=a_decimal(request.POST.get('bono_antiguedad')),
            bono_ventas=a_decimal(request.POST.get('bono_ventas')),
            comisiones_certificados=a_decimal(request.POST.get('comisiones_certificados')),
            bono_consultora=a_decimal(request.POST.get('bono_consultora')),
            otros_bonos=a_decimal(request.POST.get('otros_bonos')),
            
            aportes_afp=a_decimal(request.POST.get('aportes_afp')),
            rc_iva=a_decimal(request.POST.get('rc_iva')),
            anticipos=a_decimal(request.POST.get('anticipos')),
            prestamos=a_decimal(request.POST.get('prestamos')),
            multas=a_decimal(request.POST.get('multas')),
            rendicion_cuentas=a_decimal(request.POST.get('rendicion_cuentas')),
            pasanaku=a_decimal(request.POST.get('pasanaku')),
            
            cuenta_origen_id=cuenta_origen_id
        )
        pago.save()
        
        messages.success(request, 'El registro de planilla fue creado y los descuentos se aplicaron correctamente.')
        
        url_destino = reverse('planillas') + f'?imprimir={pago.id}'
        return redirect(url_destino)
        
    empleados = Empleado.objects.all().order_by('nombre_completo')
    cuentas = CuentaCaja.objects.all().order_by('codigo')
    return render(request, 'crear_pago.html', {'empleados': empleados, 'cuentas': cuentas})
@login_required
def detalle_curso(request, curso_id):
   

    # 1. Buscamos el curso y calculamos los totales por modalidad
    curso = get_object_or_404(Curso, id=curso_id)
    total_virtual = Inscripcion.objects.filter(curso=curso, modalidad='VIRTUAL').count()
    total_presencial = Inscripcion.objects.filter(curso=curso, modalidad='PRESENCIAL').count()
    
    # 2. Control de la pestaña activa (Virtual o Presencial)
    modalidad_param = request.GET.get('modalidad')
    if modalidad_param:
        modalidad_actual = modalidad_param.upper()
    else:
        if total_virtual > 0:
            modalidad_actual = 'VIRTUAL'
        elif total_presencial > 0:
            modalidad_actual = 'PRESENCIAL'
        else:
            modalidad_actual = 'VIRTUAL'
            
    inscritos = Inscripcion.objects.filter(curso=curso, modalidad=modalidad_actual).select_related('participante')
    
    # ==============================================================
    # --- NUEVO: BUSCADOR POR NOMBRE DE ALUMNO ---
    # ==============================================================
    buscar = request.GET.get('buscar', '').strip()
    if buscar:
        inscritos = inscritos.filter(participante__nombre_completo__icontains=buscar)
    # ==============================================================

    # --- 3. MOTOR INTELIGENTE: FILTRAR DÍAS EXACTOS DE CLASE ---
    fechas_curso = []
    
    # Quitamos la obligación de que "dias" esté lleno, con solo tener las fechas funcionará
    if curso.fecha_inicio and curso.fecha_finalizacion:
        dias_validos = set()
        # Si no escribieron días, lo tomamos como texto vacío para que no dé error
        dias_texto = str(curso.dias).upper() if curso.dias else ""
        
        if "LUNES A JUEVES" in dias_texto:
            dias_validos = {0, 1, 2, 3}
        elif "LUNES A MIERCOLES" in dias_texto or "LUNES A MIÉRCOLES" in dias_texto:
            dias_validos = {0, 1, 2}
        elif "LUNES A VIERNES" in dias_texto:
            dias_validos = {0, 1, 2, 3, 4}
        elif "FIN DE SEMANA" in dias_texto:
            dias_validos = {5, 6}
        elif "SABADO" in dias_texto or "SÁBADO" in dias_texto:
            dias_validos = {5}
        else:
            # Búsqueda de días sueltos
            if "LUNES" in dias_texto: dias_validos.add(0)
            if "MARTES" in dias_texto: dias_validos.add(1)
            if "MIERCOLES" in dias_texto or "MIÉRCOLES" in dias_texto: dias_validos.add(2)
            if "JUEVES" in dias_texto: dias_validos.add(3)
            if "VIERNES" in dias_texto: dias_validos.add(4)
            if "SABADO" in dias_texto or "SÁBADO" in dias_texto: dias_validos.add(5)
            if "DOMINGO" in dias_texto: dias_validos.add(6)
            
        # Si escribieron algo raro o lo dejaron en blanco, activamos de Lunes a Sábado por defecto
        if not dias_validos:
            dias_validos = {0, 1, 2, 3, 4, 5}

        delta = curso.fecha_finalizacion - curso.fecha_inicio
        
        # --- NUEVO: LISTADO DE FERIADOS EN PYTHON (AÑO 2026) ---
        feriados_bolivia = [
            date(2026, 1, 1), date(2026, 1, 22), date(2026, 2, 16), date(2026, 2, 17),
            date(2026, 4, 3), date(2026, 5, 1), date(2026, 6, 4), date(2026, 6, 22),
            date(2026, 8, 6), date(2026, 11, 2), date(2026, 12, 25)
        ]

        # Medida de seguridad: verificamos que la fecha final no sea antes que la de inicio
        if delta.days >= 0:
            for i in range(delta.days + 1):
                dia = curso.fecha_inicio + timedelta(days=i)
                
                # ¡CONDICIÓN ACTUALIZADA!: El día de semana coincide Y NO ES FERIADO
                if dia.weekday() in dias_validos and dia not in feriados_bolivia:
                    fechas_curso.append(dia)
                
    # --- 4. GENERACIÓN DE LA MATRIZ DE ASISTENCIA REAL (¡OPTIMIZADA!) ---
    alumnos_con_asistencia = []
    
    if inscritos and fechas_curso:
        # Extraemos solo los IDs de los alumnos para buscar rápido
        inscripciones_ids = [ins.id for ins in inscritos]
        
        # 1. Traemos TODAS las asistencias existentes de un solo golpe (1 sola consulta SQL)
        asistencias_existentes = Asistencia.objects.filter(
            inscripcion_id__in=inscripciones_ids,
            fecha__in=fechas_curso
        )
        
        # 2. Las guardamos en un diccionario rápido en memoria para buscarlas al instante
        # Clave: (inscripcion_id, fecha) -> Valor: Objeto Asistencia
        mapa_asistencias = {(a.inscripcion_id, a.fecha): a for a in asistencias_existentes}
        
        # 3. Preparamos una lista para guardar los cuadritos vacíos que falten
        nuevas_asistencias = []
        
        for inscrito in inscritos:
            for fecha in fechas_curso:
                clave = (inscrito.id, fecha)
                if clave not in mapa_asistencias:
                    # Si no existe, preparamos el objeto en RAM (PERO NO LO GUARDAMOS TODAVÍA)
                    nueva_asis = Asistencia(
                        inscripcion=inscrito, 
                        fecha=fecha, 
                        estado='PENDIENTE'
                    )
                    nuevas_asistencias.append(nueva_asis)
        
        # 4. GUARDADO MASIVO: Guarda miles de registros en 1 sola consulta SQL
        if nuevas_asistencias:
            Asistencia.objects.bulk_create(nuevas_asistencias)
            
            # Recargamos el diccionario de forma segura para tener los IDs reales de la base de datos
            asistencias_existentes = Asistencia.objects.filter(
                inscripcion_id__in=inscripciones_ids,
                fecha__in=fechas_curso
            )
            mapa_asistencias = {(a.inscripcion_id, a.fecha): a for a in asistencias_existentes}

        # 5. Finalmente, armamos la matriz para el HTML usando solo la memoria RAM (ultra rápido)
        for inscrito in inscritos:
            asistencias_alumno = []
            for fecha in fechas_curso:
                asistencias_alumno.append(mapa_asistencias[(inscrito.id, fecha)])
            
            alumnos_con_asistencia.append({
                'inscripcion': inscrito,
                'asistencias': asistencias_alumno
            })
    
    contexto = {
        'curso': curso,
        'inscritos': inscritos,
        'alumnos_con_asistencia': alumnos_con_asistencia,
        'fechas_curso': fechas_curso,
        'modalidad_actual': modalidad_actual,
        'total_virtual': total_virtual,
        'total_presencial': total_presencial,
        'buscar': buscar,
    }
    return render(request, 'detalle_curso.html', contexto)
@csrf_exempt
@login_required
def toggle_asistencia(request):
    if request.method == 'POST':
        # Solo dejamos la lógica original que lee JSON
        data = json.loads(request.body)
        asistencia_id = data.get('id')
        asistencia = get_object_or_404(Asistencia, id=asistencia_id)
        
        # Máquina rotativa: PENDIENTE -> PRESENTE -> FALTA -> PENDIENTE
        if asistencia.estado == 'PENDIENTE':
            asistencia.estado = 'PRESENTE'
        elif asistencia.estado == 'PRESENTE':
            asistencia.estado = 'FALTA'
        else:
            asistencia.estado = 'PENDIENTE'
            
        asistencia.save()
        return JsonResponse({'status': 'ok', 'nuevo_estado': asistencia.estado})

@login_required
def descargar_pdf_asistencia(request, curso_id):
    # 1. Traemos los datos del curso y la modalidad actual
    curso = get_object_or_404(Curso, id=curso_id)
    modalidad_actual = request.GET.get('modalidad', 'VIRTUAL')
    
    # 2. Traemos a los alumnos correspondientes
    inscritos = Inscripcion.objects.filter(curso=curso, modalidad=modalidad_actual).select_related('participante')
    
    # 3. Preparamos los datos para enviarlos a la plantilla PDF
    contexto = {
        'curso': curso,
        'inscritos': inscritos,
        'modalidad_actual': modalidad_actual,
    }
    
    # 4. Enlazamos con la plantilla HTML exclusiva para el PDF
    template = get_template('pdf_asistencia.html')
    html = template.render(contexto)
    
    # 5. Creamos la respuesta en formato PDF
    response = HttpResponse(content_type='application/pdf')
    # Quita la palabra 'attachment;' si quieres que el PDF se abra en una nueva pestaña en lugar de descargarse directo
    response['Content-Disposition'] = f'attachment; filename="Lista_{curso.nombre}_{modalidad_actual}.pdf"'
    
    # Renderizamos el PDF (Esta línea asume que usas xhtml2pdf)
    from xhtml2pdf import pisa
    pisa_status = pisa.CreatePDF(html, dest=response)
    
    if pisa_status.err:
        return HttpResponse('Hubo un error al generar el PDF')
    
    return response

@login_required
@user_passes_test(es_administrador)
def registrar_anticipo(request, honorario_id):
    # Buscamos el honorario específico usando su ID
    honorario = get_object_or_404(Honorario, id=honorario_id)
    
    if request.method == 'POST':
        monto = request.POST.get('monto_anticipo')
        detalle = request.POST.get('detalle_anticipo')
        cuenta_id = request.POST.get('cuenta_origen_id') # <-- NUEVO: Atrapamos la cuenta elegida
        
        if monto:
            monto_float = float(monto)
            
            # Validación de seguridad contable (Límite de monto)
            if monto_float > honorario.monto_acordado:
                messages.error(request, f'No se pudo realizar la acción. El anticipo de Bs. {monto_float} supera el límite del honorario total (Bs. {honorario.monto_acordado}).')
                return redirect('registrar_anticipo', honorario_id=honorario.id)
            
            # Guardamos el anticipo en el Honorario
            honorario.anticipo = monto_float 
            honorario.fecha_anticipo = date.today() 
            if detalle:
                honorario.observacion = detalle 
            honorario.save() 
            
            # --- NUEVO: AUTOMATIZACIÓN EN EL FLUJO DE CAJA ---
            if cuenta_id:
                try:
                    cuenta_origen = CuentaCaja.objects.get(id=cuenta_id)
                    detalle_movimiento = f"Anticipo Honorario: {honorario.curso.docente.nombre} - Curso: {honorario.curso.nombre}"
                    
                    # Buscamos si ya existía un movimiento de este anticipo para no duplicarlo si lo editan
                    movimiento_existente = MovimientoCaja.objects.filter(
                        detalle=detalle_movimiento,
                        tipo='SALIDA'
                    ).first()
                    
                    if movimiento_existente:
                        movimiento_existente.monto = monto_float
                        movimiento_existente.cuenta = cuenta_origen
                        movimiento_existente.fecha = honorario.fecha_anticipo
                        movimiento_existente.save()
                    else:
                        MovimientoCaja.objects.create(
                            fecha=honorario.fecha_anticipo,
                            detalle=detalle_movimiento,
                            cuenta=cuenta_origen,
                            tipo='SALIDA',
                            monto=monto_float
                        )
                except Exception as e:
                    print(f"Error al registrar anticipo en caja: {e}")
            
            messages.success(request, f'Anticipo de Bs. {monto_float} registrado y debitado de la cuenta exitosamente.')
            return redirect('honorarios') 
            
    # GET: Enviamos el catálogo de cuentas ordenado (001, 002...)
    cuentas = CuentaCaja.objects.all().order_by('codigo')
    return render(request, 'registrar_anticipo.html', {'honorario': honorario, 'cuentas': cuentas})

@login_required
@user_passes_test(es_administrador)
def pagar_honorario(request, honorario_id):
    # 1. Buscamos el honorario correspondiente
    honorario = get_object_or_404(Honorario, id=honorario_id)
    
    # 2. FLUJO DE PROCESAMIENTO / CONFIRMACIÓN
    if honorario.estado != 'PAGADO':
        if request.method == 'POST':
            # Capturamos la cuenta seleccionada en el menú desplegable del HTML
            cuenta_id = request.POST.get('cuenta_origen_id')
            
            # Actualizamos el estado a PAGADO
            honorario.estado = 'PAGADO'
            honorario.fecha_pago = date.today()
            honorario.save()
            
            # Automatización del registro en el Flujo de Caja con la cuenta REAL
            try:
                # Buscamos la cuenta exacta que el usuario seleccionó (Ej: 002 - BANCO)
                cuenta_seleccionada = CuentaCaja.objects.get(id=cuenta_id)
                
                MovimientoCaja.objects.create(
                    fecha=honorario.fecha_pago,
                    detalle=f"Pago de Honorario: {honorario.curso.docente.nombre} - Curso: {honorario.curso.nombre}",
                    cuenta=cuenta_seleccionada, # Asignamos la cuenta elegida
                    tipo='SALIDA',
                    monto=honorario.saldo
                )
            except Exception as e:
                print(f"Error al registrar movimiento en caja: {e}")
            
            # Pasamos directo a generar el PDF
            pass  
        else:
            # GET: Traemos todas tus cuentas (001, 002, 003...) ordenadas para mostrarlas
            cuentas_disponibles = CuentaCaja.objects.all().order_by('codigo')
            return render(request, 'confirmar_pago.html', {
                'honorario': honorario,
                'cuentas': cuentas_disponibles # Enviamos las cuentas a la pantalla
            })
        
    # 3. FLUJO DE GENERACIÓN E IMPRESIÓN DEL PDF (Esto se mantiene intacto)
    response = HttpResponse(content_type='application/pdf')
    nombre_archivo = f"Boleta_Honorario_{honorario.id}.pdf"
    response['Content-Disposition'] = f'inline; filename="{nombre_archivo}"'
    
    # DIBUJAR LA BOLETA 
    p = canvas.Canvas(response, pagesize=letter)
    width, height = letter
    
    p.setFillColor(colors.HexColor("#1e3a8a"))
    p.rect(0, height - 80, width, 80, fill=True, stroke=False)
    
    p.setFillColor(colors.white)
    p.setFont("Helvetica-Bold", 22)
    p.drawString(40, height - 48, "LIMA S.R.L.")
    p.setFont("Helvetica", 10)
    p.drawString(40, height - 64, "Gestión Académica Profesional")
    
    p.drawRightString(width - 40, height - 42, f"Fecha de Pago: {honorario.fecha_pago.strftime('%d/%m/%Y') if honorario.fecha_pago else date.today().strftime('%d/%m/%Y')}")
    p.setFont("Helvetica-Bold", 12)
    p.drawRightString(width - 40, height - 62, f"COMPROBANTE Nº: H-{honorario.id:04d}")
    
    p.setFillColor(colors.HexColor("#1f2937"))
    p.setFont("Helvetica-Bold", 15)
    p.drawString(40, height - 130, "BOLETA DE LIQUIDACIÓN DE HONORARIOS DOCENTES")
    
    p.setStrokeColor(colors.HexColor("#d1d5db"))
    p.setLineWidth(1)
    p.line(40, height - 140, width - 40, height - 140)
    
    p.setFont("Helvetica-Bold", 11)
    p.drawString(40, height - 175, "Nombre del Docente:")
    p.setFont("Helvetica", 11)
    p.drawString(170, height - 175, str(honorario.curso.docente.nombre))
    
    p.setFont("Helvetica-Bold", 11)
    p.drawString(40, height - 200, "Curso Dictado:")
    p.setFont("Helvetica", 11)
    p.drawString(170, height - 200, str(honorario.curso.nombre))
    
    p.setFont("Helvetica-Bold", 11)
    p.drawString(40, height - 225, "Carga Realizada:")
    p.setFont("Helvetica", 11)
    p.drawString(170, height - 225, str(honorario.carga))
    
    p.setFillColor(colors.HexColor("#f9fafb"))
    p.rect(40, height - 350, width - 80, 100, fill=True, stroke=True)
    
    p.setFillColor(colors.HexColor("#374151"))
    p.setFont("Helvetica", 11)
    p.drawString(60, height - 280, "Honorario Base Acordado:")
    p.drawRightString(width - 60, height - 280, f"Bs. {honorario.honorario_total:.2f}")
    
    p.drawString(60, height - 305, "Menos Anticipos Registrados:")
    p.setFillColor(colors.HexColor("#b91c1c"))
    p.drawRightString(width - 60, height - 305, f"- Bs. {honorario.anticipo:.2f}")
    
    p.setStrokeColor(colors.HexColor("#e5e7eb"))
    p.line(60, height - 315, width - 60, height - 315)
    
    p.setFillColor(colors.HexColor("#1f2937"))
    p.setFont("Helvetica-Bold", 12)
    p.drawString(60, height - 335, "LIQUIDACIÓN NETO RECIBIDO:")
    p.setFillColor(colors.HexColor("#16a34a"))
    p.drawRightString(width - 60, height - 335, f"Bs. {honorario.saldo:.2f}")
    
    p.setFillColor(colors.HexColor("#4b5563"))
    p.setFont("Helvetica-Bold", 10)
    p.drawString(40, height - 380, "Glosa / Observación:")
    p.setFont("Helvetica-Oblique", 10)
    p.drawString(40, height - 398, str(honorario.observacion if honorario.observacion else "Liquidación total de haberes sin observaciones adicionales."))
    
    p.setStrokeColor(colors.HexColor("#9ca3af"))
    p.line(70, height - 500, 230, height - 500)
    p.setFont("Helvetica", 9)
    p.setFillColor(colors.HexColor("#374151"))
    p.drawCentredString(150, height - 515, "Firma del Docente Receptor")
    
    p.line(width - 230, height - 500, width - 70, height - 500)
    p.drawCentredString(width - 150, height - 515, "Administración / LIMA S.R.L.")
    
    p.showPage()
    p.save()
    
    return response

@login_required
@user_passes_test(es_administrador)
def editar_movimiento(request, movimiento_id):
    movimiento = get_object_or_404(MovimientoCaja, id=movimiento_id)
    cuentas = CuentaCaja.objects.all().order_by('codigo')

    if request.method == 'POST':
        # --- CASO A: ELIMINAR REGISTRO ---
        if 'eliminar' in request.POST:
            movimiento.delete()
            # AGREGAMOS EL MENSAJE AQUÍ:
            messages.success(request, 'El registro contable fue eliminado correctamente.')
            return redirect('flujo_caja')

        # --- CASO B: EDITAR REGISTRO ---
        movimiento.fecha = request.POST.get('fecha')
        movimiento.detalle = request.POST.get('detalle')
        movimiento.tipo = request.POST.get('tipo')
        movimiento.monto = request.POST.get('monto')
        
        cuenta_id = request.POST.get('cuenta_id')
        if cuenta_id:
            movimiento.cuenta = CuentaCaja.objects.get(id=cuenta_id)
            
        movimiento.save()
        # AGREGAMOS EL MENSAJE AQUÍ:
        messages.success(request, 'El registro contable fue actualizado con éxito.')
        return redirect('flujo_caja')

    return render(request, 'editar_movimiento.html', {
        'movimiento': movimiento,
        'cuentas': cuentas
    })

@login_required
@user_passes_test(es_administrador)
def editar_servicio(request, servicio_id):
    # 1. Buscamos el servicio exacto
    servicio = get_object_or_404(ServicioConsultora, id=servicio_id)

    if request.method == 'POST':
        # --- CASO A: EL USUARIO PRESIONÓ "ELIMINAR REGISTRO" ---
        if 'eliminar' in request.POST:
            servicio.delete()
            # ¡NUEVO! Dispara la ventanita roja
            messages.success(request, 'El registro fue eliminado y el dinero retornó al flujo de caja correctamente.')
            return redirect('consultora')

        # --- CASO B: EL USUARIO PRESIONÓ "HECHO" (GUARDAR CAMBIOS) ---
        form = ServicioConsultoraForm(request.POST, request.FILES, instance=servicio)
        if form.is_valid():
            servicio_guardado = form.save(commit=False)
            
            # Recalculamos el 20% por si el usuario modificó el importe
            if servicio_guardado.importe:
                servicio_guardado.comision = servicio_guardado.importe * Decimal('0.20')
            else:
                servicio_guardado.comision = Decimal('0.00')
                
            servicio_guardado.save()
            # ¡NUEVO! Dispara la ventanita azul
            messages.success(request, 'El trámite contable fue actualizado con éxito.')
            return redirect('consultora')
    else:
        # Cargamos el formulario con los datos actuales del servicio
        form = ServicioConsultoraForm(instance=servicio)

    return render(request, 'editar_servicio.html', {
        'form': form,
        'servicio': servicio
    })

@login_required
def portal_inicio(request):
    return render(request, 'portal.html')

def editar_inscripcion(request, id):
    # 1. Buscamos el registro exacto en la base de datos
    inscripcion = get_object_or_404(Inscripcion, id=id)
    cursos = Curso.objects.all()


    if request.method == 'POST':
        # 2. Extraemos los nuevos datos enviados por el formulario
        nombre_completo = request.POST.get('nombre_completo')
        celular = request.POST.get('celular')
        curso_id = request.POST.get('curso')
        modalidad = request.POST.get('modalidad')
        importe = request.POST.get('importe')
        inscripcion.saldo_pendiente = Decimal(str(request.POST.get('saldo_pendiente', '0.00') or '0.00'))
        inscripcion.save()
        forma_pago = request.POST.get('forma_pago')
        banco = request.POST.get('banco')
        vendedor = request.POST.get('vendedor').upper()
        
        # --- AQUÍ ATRAPAMOS EL NUEVO CAMPO ---
        registrado_por = request.POST.get('registrado_por').upper()
        
        fecha_inscripcion = request.POST.get('fecha_inscripcion')

        # Actualizar el participante vinculado
        participante = inscripcion.participante
        if hasattr(participante, 'nombre_completo'):
            participante.nombre_completo = nombre_completo
        else:
            participante.nombre = nombre_completo
            
        participante.celular = celular
        participante.save() # Guarda los cambios del alumno

        # Actualizar datos principales de la inscripción
        inscripcion.curso = get_object_or_404(Curso, id=curso_id)
        inscripcion.modalidad = modalidad
            
        inscripcion.modalidad = modalidad
        inscripcion.importe = importe
        inscripcion.forma_pago = forma_pago
        inscripcion.banco = banco
        inscripcion.vendedor = vendedor
        
        # --- AQUÍ LO GUARDAMOS EN LA INSCRIPCIÓN ---
        inscripcion.registrado_por = registrado_por
        
        inscripcion.fecha_inscripcion = fecha_inscripcion
        
        # Guardamos los cambios definitivos
        inscripcion.save()
        
        # --- NUEVO: MENSAJE QUE SE CIERRA SOLO ---
        messages.success(request, 'Inscripción actualizada con éxito.')
        
        return redirect('inscripciones')
    hoy = date.today()
    
    # Aplicamos el mismo orden inteligente: 1 (No inició), 2 (En curso), 3 (Finalizado)
    cursos = Curso.objects.select_related('docente').annotate(
        orden_estado=Case(
            When(fecha_inicio__gt=hoy, then=Value(1)),
            When(fecha_inicio__lte=hoy, fecha_finalizacion__gte=hoy, then=Value(2)),
            default=Value(3),
            output_field=IntegerField(),
        )
    ).order_by('orden_estado', '-fecha_inicio')
    
    context = {
        'inscripcion': inscripcion,
        'cursos': cursos,
        'hoy': hoy, # ¡Importante pasar hoy para que el JS sepa clasificar!
    }
    return render(request, 'editar_inscripcion.html', context)

@login_required
def editar_inscripcion_cc(request, id):
    from decimal import Decimal
    from django.db.models import Case, When, Value, IntegerField
    
    # Buscamos el registro exacto
    inscripcion = get_object_or_404(Inscripcion, id=id)

    if request.method == 'POST':
        # 1. Atrapamos los datos del formulario
        nombre_completo = request.POST.get('nombre_completo', '').strip().upper()
        celular = request.POST.get('celular', '').strip()
        curso_id = request.POST.get('curso')
        fecha = request.POST.get('fecha_inscripcion')
        
        monto_pagado = Decimal(str(request.POST.get('monto_pagado', '0.00').strip() or '0.00'))
        saldo_pendiente = Decimal(str(request.POST.get('saldo_pendiente', '0.00').strip() or '0.00'))
        
        banco = request.POST.get('banco')
        forma_pago = request.POST.get('forma_pago')
        modalidad = request.POST.get('modalidad', 'VIRTUAL').upper()
        vendedor = request.POST.get('vendedor', '').upper()
        registrado_por = request.POST.get('registrado_por', '').upper()

        # 2. Actualizar Participante
        participante = inscripcion.participante
        if hasattr(participante, 'nombre_completo'):
            participante.nombre_completo = nombre_completo
        else:
            participante.nombre = nombre_completo
        participante.celular = celular
        participante.save()

        # 3. Actualizar Inscripción
        inscripcion.curso = get_object_or_404(Curso, id=curso_id)
        inscripcion.modalidad = modalidad
        inscripcion.importe = monto_pagado
        inscripcion.saldo_pendiente = saldo_pendiente
        inscripcion.banco = banco
        inscripcion.forma_pago = forma_pago
        inscripcion.vendedor = vendedor
        inscripcion.registrado_por = registrado_por
        inscripcion.fecha_inscripcion = fecha
        
        inscripcion.save()
        
        # 🚀 REDIRECCIÓN FORZADA A LA TABLA CORRECTA
        messages.success(request, 'La deuda del alumno fue actualizada con éxito.')
        return redirect('cuentas_por_cobrar') 
        
    hoy = date.today()
    cursos = Curso.objects.select_related('docente').annotate(
        orden_estado=Case(
            When(fecha_inicio__gt=hoy, then=Value(1)),
            When(fecha_inicio__lte=hoy, fecha_finalizacion__gte=hoy, then=Value(2)),
            default=Value(3),
            output_field=IntegerField(),
        )
    ).order_by('orden_estado', '-fecha_inicio')
    
    return render(request, 'editar_inscripcion_cc.html', {
        'inscripcion': inscripcion,
        'cursos': cursos
    })

def eliminar_inscripcion(request, id):
    # 1. Buscamos el registro exacto
    inscripcion = get_object_or_404(Inscripcion, id=id)
    
    # 2. Eliminamos el registro
    inscripcion.delete()
    
    # 3. Creamos el mensaje de éxito (NUEVO)
    messages.success(request, 'El registro de inscripción fue eliminado correctamente.')
    
    # 4. Recargamos la tabla general
    return redirect('inscripciones')

@login_required
def crear_venta_servicio(request):
    if request.method == 'POST':
        # 1. Atrapamos los datos del cliente
        nombre_completo = request.POST.get('nombre_completo').upper()
        celular = request.POST.get('celular')
        
        # 2. Buscamos si el participante ya existe (por celular) para no duplicarlo
        participantes = Participante.objects.filter(celular=celular)
        if participantes.exists():
            participante = participantes.first()
            # Actualizamos su nombre por si hubo algún cambio
            if hasattr(participante, 'nombre_completo'):
                participante.nombre_completo = nombre_completo
            else:
                participante.nombre = nombre_completo
            participante.save()
        else:
            # Si es un cliente nuevo, lo creamos
            try:
                participante = Participante.objects.create(nombre_completo=nombre_completo, celular=celular)
            except:
                participante = Participante.objects.create(nombre=nombre_completo, celular=celular)

        # 3. Atrapamos los datos financieros y del servicio
        tipo_servicio = request.POST.get('tipo_servicio')
        detalle = request.POST.get('detalle').upper()
        importe = request.POST.get('importe')
        forma_pago = request.POST.get('forma_pago')
        banco = request.POST.get('banco')
        vendedor = request.POST.get('vendedor').upper()
        registrado_por = request.POST.get('registrado_por').upper()
        fecha_venta = request.POST.get('fecha_venta')

        # 4. Guardamos la Venta Oficial en la nueva tabla
        VentaServicio.objects.create(
            participante=participante,
            tipo_servicio=tipo_servicio,
            detalle=detalle,
            importe=importe,
            forma_pago=forma_pago,
            banco=banco,
            vendedor=vendedor,
            registrado_por=registrado_por,
            fecha_venta=fecha_venta
        )
        
        # Volvemos a la Súper Tabla de Inscripciones donde está todo unificado
        return redirect('inscripciones')

    return render(request, 'crear_venta_servicio.html')

@login_required
def editar_venta_servicio(request, id):
    venta = get_object_or_404(VentaServicio, id=id)
    
    if request.method == 'POST':
        # 1. Actualizamos los datos del participante vinculado
        participante = venta.participante
        participante.nombre_completo = request.POST.get('nombre_completo').upper()
        participante.celular = request.POST.get('celular')
        participante.save()
        
        # 2. Actualizamos los datos comerciales
        venta.tipo_servicio = request.POST.get('tipo_servicio')
        detalle_raw = request.POST.get('detalle')
        venta.detalle = detalle_raw.upper() if detalle_raw else None
        venta.importe = request.POST.get('importe')
        venta.forma_pago = request.POST.get('forma_pago')
        venta.banco = request.POST.get('banco')
        venta.vendedor = request.POST.get('vendedor').upper()
        venta.registrado_por = request.POST.get('registrado_por').upper()
        venta.fecha_venta = request.POST.get('fecha_venta')
        
        venta.save()
        
        messages.success(request, f'Venta de {venta.tipo_servicio} actualizada con éxito.')
        return redirect('inscripciones')
        
    return render(request, 'editar_venta_servicio.html', {'venta': venta})

@login_required
def eliminar_venta_servicio(request, id):
    venta = get_object_or_404(VentaServicio, id=id)
    tipo = venta.tipo_servicio
    venta.delete()
    
    messages.success(request, f'La venta de {tipo} fue eliminada correctamente.')
    return redirect('inscripciones')

from django.shortcuts import get_object_or_404, redirect

@login_required
def generar_certificados_curso(request, curso_id):
    curso = get_object_or_404(Curso, id=curso_id)
    
    modalidad_actual = request.GET.get('modalidad', 'VIRTUAL')
    inscritos = Inscripcion.objects.filter(curso=curso, modalidad=modalidad_actual).select_related('participante')
    
    # 1. Selector Inteligente por Docente
    nombre_docente = curso.docente.nombre.lower()
    
    if 'juan' in nombre_docente or 'juanjo' in nombre_docente:
        archivo_fondo = 'certificado_juanjo.jpg'
    elif 'mariana' in nombre_docente:
        archivo_fondo = 'certificado_mariana.jpg'
    elif 'rodrigo' in nombre_docente:
        archivo_fondo = 'certificado_rodrigo.jpg'
    else:
        archivo_fondo = 'certificado.jpg'

    ruta_actual = os.path.dirname(os.path.abspath(__file__))
    ruta_imagen = os.path.join(ruta_actual, 'static', archivo_fondo).replace('\\', '/')
    ruta_fuente = os.path.join(ruta_actual, 'static', 'Montserrat-Bold.ttf').replace('\\', '/')
    
    template = get_template('pdf_certificados.html')
    
    # 2. Creamos la carpeta ZIP virtual en la memoria
    buffer_zip = BytesIO()
    
    with zipfile.ZipFile(buffer_zip, 'w', zipfile.ZIP_DEFLATED) as archivo_zip:
        for inscrito in inscritos:
            # Preparamos los datos del alumno actual
            contexto = {
                'curso': curso,
                'inscrito': inscrito,
                'ruta_imagen': ruta_imagen,
                'ruta_fuente': ruta_fuente,
            }
            html = template.render(contexto)
            
            # Creamos el PDF en la memoria
            pdf_buffer = BytesIO()
            pisa_status = pisa.CreatePDF(html, dest=pdf_buffer)
            
            if not pisa_status.err:
                # Nombre del archivo para este alumno específico
                nombre_limpio = inscrito.participante.nombre_completo.replace(" ", "_")
                nombre_archivo = f"Certificado_{nombre_limpio}.pdf"
                
                # Metemos el PDF dentro de la carpeta ZIP
                archivo_zip.writestr(nombre_archivo, pdf_buffer.getvalue())

    # 3. Preparamos la carpeta comprimida para ser descargada
    buffer_zip.seek(0)
    response = HttpResponse(buffer_zip, content_type='application/zip')
    # "attachment" fuerza la descarga de la carpeta .zip
    response['Content-Disposition'] = f'attachment; filename="Certificados_{curso.nombre}_{modalidad_actual}.zip"'
    
    return response
@login_required
@user_passes_test(es_administrador)
def exportar_excel_caja(request):
    # 1. Creamos el archivo Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Libro Diario"
    
    # --- NUEVOS COLORES: BLANCOS, GRISES Y NEGRES ---
    header_fill = PatternFill(start_color="111827", end_color="111827", fill_type="solid") # Negro Oscuro (Slate 900)
    row_impar_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid") # Blanco
    row_par_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid") # Gris Ultra Claro
    
    font_white_bold = Font(color="FFFFFF", bold=True, size=11)
    
    align_center = Alignment(horizontal="center", vertical="center")
    align_right = Alignment(horizontal="right", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")
    
    border_thin = Border(
        left=Side(style='thin', color='D1D5DB'),
        right=Side(style='thin', color='D1D5DB'),
        top=Side(style='thin', color='D1D5DB'),
        bottom=Side(style='thin', color='D1D5DB')
    )
    
    # Escribir Cabeceras
    headers = ["FECHA", "DETALLE DE OPERACIÓN", "COD", "ENTRADAS", "SALIDAS", "SALDO"]
    ws.append(headers)
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = font_white_bold
        cell.alignment = align_center
        cell.border = border_thin
        
    ws.auto_filter.ref = "A1:F1"
    
    # --- NUEVO: ATRAPAR EL FILTRO DE FECHAS ---
    rango_fechas = request.GET.get('rango_fechas', '')
    movimientos_qs = MovimientoCaja.objects.all().order_by('fecha', 'id')
    
    if rango_fechas:
        if ' a ' in rango_fechas:
            fecha_inicio, fecha_fin = rango_fechas.split(' a ')
            movimientos_qs = movimientos_qs.filter(fecha__range=[fecha_inicio, fecha_fin])
        elif len(rango_fechas) == 4 and rango_fechas.isdigit():
            movimientos_qs = movimientos_qs.filter(fecha__year=rango_fechas)
        elif len(rango_fechas) == 7 and '-' in rango_fechas:
            anio, mes = rango_fechas.split('-')
            movimientos_qs = movimientos_qs.filter(fecha__year=anio, fecha__month=mes)
        else:
            movimientos_qs = movimientos_qs.filter(fecha=rango_fechas)
    
    # Calculadora de saldos sobre los registros YA FILTRADOS
    saldo_corriente = Decimal('0.00')
    total_entradas = Decimal('0.00')
    total_salidas = Decimal('0.00')
    
    for i, mov in enumerate(movimientos_qs, start=2):
        entrada = mov.monto if mov.tipo == 'ENTRADA' else ""
        salida = mov.monto if mov.tipo == 'SALIDA' else ""
        
        if mov.tipo == 'ENTRADA':
            saldo_corriente += mov.monto
            total_entradas += mov.monto
        else:
            saldo_corriente -= mov.monto
            total_salidas += mov.monto
            
        ws.append([mov.fecha, mov.detalle.upper(), mov.cuenta.codigo, entrada, salida, saldo_corriente])
        
        fill_actual = row_par_fill if i % 2 == 0 else row_impar_fill
        for col_num in range(1, 7):
            cell = ws.cell(row=i, column=col_num)
            cell.fill = fill_actual
            cell.border = border_thin
            if col_num == 1:
                cell.number_format = 'DD/MM/YYYY'
                cell.alignment = align_center
            elif col_num == 3:
                cell.alignment = align_center
            elif col_num in [4, 5, 6]:
                cell.number_format = '"Bs" #,##0.00'
                cell.alignment = align_right
            else:
                cell.alignment = align_left

    # Fila final de Totales
    ultima_fila = len(movimientos_qs) + 2
    ws.append(["TOTALES", f"BALANCE DEL PERIODO: {rango_fechas if rango_fechas else 'HISTÓRICO'}", "", total_entradas, total_salidas, saldo_corriente])
    for col_num in range(1, 7):
        cell = ws.cell(row=ultima_fila, column=col_num)
        cell.fill = header_fill
        cell.font = font_white_bold
        if col_num in [4, 5, 6]:
            cell.number_format = '"Bs" #,##0.00'
            cell.alignment = align_right
        else:
            cell.alignment = align_center

    # Ajustar columnas
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 20

    # Hoja de Resumen Mensual (También filtrada a Blanco y Negro)
    ws_mes = wb.create_sheet(title="Resumen Mensual")
    ws_mes.append(["MES", "TOTAL ENTRADAS", "TOTAL SALIDAS", "FLUJO NETO DEL MES"])
    for col_num in range(1, 5):
        c = ws_mes.cell(row=1, column=col_num)
        c.fill = header_fill
        c.font = font_white_bold
        c.alignment = align_center
        c.border = border_thin

    resumen_mensual = movimientos_qs.annotate(mes=TruncMonth('fecha')).values('mes').annotate(
        in_total=Sum('monto', filter=Q(tipo='ENTRADA')),
        out_total=Sum('monto', filter=Q(tipo='SALIDA'))
    ).order_by('mes')

    for i, res in enumerate(resumen_mensual, start=2):
        ent = res['in_total'] or Decimal('0.00')
        sal = res['out_total'] or Decimal('0.00')
        neto = ent - sal
        ws_mes.append([res['mes'].strftime("%B %Y").upper(), ent, sal, neto])
        
        fill_m = row_par_fill if i % 2 == 0 else row_impar_fill
        for col_num in range(1, 5):
            c = ws_mes.cell(row=i, column=col_num)
            c.fill = fill_m
            c.border = border_thin
            if col_num == 1:
                c.alignment = align_center
            else:
                c.number_format = '"Bs" #,##0.00'
                c.alignment = align_right

    ws_mes.column_dimensions['A'].width = 25
    ws_mes.column_dimensions['B'].width = 20
    ws_mes.column_dimensions['C'].width = 20
    ws_mes.column_dimensions['D'].width = 25

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    texto_archivo = rango_fechas.replace(" ", "") if rango_fechas else "Completo"
    response['Content-Disposition'] = f'attachment; filename="Reporte_Caja_{texto_archivo}.xlsx"'
    wb.save(response)
    
    return response

@login_required
def obtener_datos_empleado_pago(request, empleado_id):
    empleado = get_object_or_404(Empleado, id=empleado_id)
    mes_str = request.GET.get('mes', '') 
    
    datos = {
        'salario_base': float(empleado.salario_base),
        'bono_antiguedad': 0.00,
        'bono_ventas': 0.00,
        'comisiones_certificados': 0.00,
        'bono_consultora': 0.00,
        'descuento_prestamo': 0.00,
        'descuento_anticipo': 0.00 
    }

    if not mes_str:
        return JsonResponse(datos)

    try:
        from decimal import Decimal
        anio, mes = map(int, mes_str.split('-'))
        nombre_completo_emp = empleado.nombre_completo.lower()
        
        # 1. Bono de Antigüedad
        if empleado.fecha_ingreso:
            fecha_evaluacion = date(anio, mes, 1)
            dias_antiguedad = (fecha_evaluacion - empleado.fecha_ingreso).days
            if dias_antiguedad >= 730: 
                datos['bono_antiguedad'] = 165.00 

        # =========================================================
        # 2. BONO DE VENTAS (Cursos + Grabaciones + Sistemas + Otros)
        # =========================================================
        # A) Sumamos las inscripciones normales a cursos
        inscripciones_mes = Inscripcion.objects.filter(fecha_inscripcion__year=anio, fecha_inscripcion__month=mes)
        total_insc = sum(insc.importe for insc in inscripciones_mes if insc.vendedor and insc.vendedor.lower() in nombre_completo_emp)
        
        # B) Sumamos las ventas de Grabaciones, Sistemas y Otros
        ventas_extra = VentaServicio.objects.filter(
            fecha_venta__year=anio, 
            fecha_venta__month=mes,
            tipo_servicio__in=['GRABACIÓN', 'SISTEMA', 'OTRO']
        )
        total_extra = sum(venta.importe for venta in ventas_extra if venta.vendedor and venta.vendedor.lower() in nombre_completo_emp)
        
        # C) Calculamos el 10% del total combinado
        datos['bono_ventas'] = float((Decimal(total_insc) + Decimal(total_extra)) * Decimal('0.10')) 

        # =========================================================
        # 3. Comisiones por Certificados (Ventas Individuales)
        # =========================================================
        ventas_certificados = VentaServicio.objects.filter(tipo_servicio='CERTIFICADO', fecha_venta__year=anio, fecha_venta__month=mes)
        total_cert = sum(venta.importe for venta in ventas_certificados if venta.vendedor and venta.vendedor.lower() in nombre_completo_emp)
        comision_individual = float(Decimal(total_cert) * Decimal('0.10'))
        
        # Bono por envío masivo de certificados de cursos (Ej. para Patricia)
        bono_envio_lotes = 0.00
        if 'patricia' in nombre_completo_emp:
            cursos_enviados = Curso.objects.filter(certificados_enviados=True, fecha_envio_certificados__year=anio, fecha_envio_certificados__month=mes)
            for curso_env in cursos_enviados:
                total_alumnos_curso = Inscripcion.objects.filter(curso=curso_env).count()
                bono_envio_lotes += (total_alumnos_curso * 1.00)
                
        datos['comisiones_certificados'] = comision_individual + float(bono_envio_lotes)
        
        # 4. Bono Consultora
        servicios_consultora = ServicioConsultora.objects.filter(fecha__year=anio, fecha__month=mes)
        total_consultora = Decimal('0.00')
        for servicio in servicios_consultora:
            if servicio.contador:
                nombre_contador = servicio.contador.nombre_completo.lower().strip()
                if nombre_contador == 'juanjo' and 'juan jose' in nombre_completo_emp:
                    total_consultora += servicio.comision
                elif nombre_contador in nombre_completo_emp:
                    total_consultora += servicio.comision
        datos['bono_consultora'] = float(total_consultora)

        # 5. ESCANEO Y DESCUENTO AUTOMÁTICO DE PRÉSTAMOS
        prestamos_activos = Prestamo.objects.filter(empleado=empleado, estado='ACTIVO')
        total_descuento = Decimal('0.00')
        for prestamo in prestamos_activos:
            cuota_sugerida = prestamo.total_deuda / Decimal(prestamo.nro_cuotas)
            cuota_a_cobrar = min(cuota_sugerida, prestamo.saldo_restante)
            total_descuento += cuota_a_cobrar
            
        datos['descuento_prestamo'] = float(total_descuento)

        # 6. ESCANEO Y DESCUENTO AUTOMÁTICO DE ANTICIPOS (A prueba de errores)
        meses_dict = {
            1: 'ENERO', 2: 'FEBRERO', 3: 'MARZO', 4: 'ABRIL',
            5: 'MAYO', 6: 'JUNIO', 7: 'JULIO', 8: 'AGOSTO',
            9: 'SEPTIEMBRE', 10: 'OCTUBRE', 11: 'NOVIEMBRE', 12: 'DICIEMBRE'
        }
        mes_literal = f"{meses_dict.get(mes, '')} {anio}" # Crea "JULIO 2026"
        
        from django.db.models import Q
        anticipos_activos = AnticipoEmpleado.objects.filter(
            empleado=empleado, 
            estado='PENDIENTE'
        ).filter(
            Q(mes_descuento__icontains=mes_str) | Q(mes_descuento__icontains=mes_literal)
        )
        
        total_anticipos = sum(ant.monto for ant in anticipos_activos)
        datos['descuento_anticipo'] = float(total_anticipos)

    except Exception as e:
        print(f"Error en automatización de sueldos y préstamos: {e}")

    return JsonResponse(datos)

@login_required
@user_passes_test(es_administrador)
def eliminar_pago(request, pago_id):
    pago = get_object_or_404(PagoSueldo, id=pago_id)
    pago.delete() # Esto activará la reversión automática que acabamos de programar
    
    messages.success(request, 'El registro de pago fue eliminado y el dinero retornó al flujo de caja correctamente.')
    return redirect('planillas')

@login_required
@user_passes_test(es_administrador)
def editar_pago(request, pago_id):
    pago = get_object_or_404(PagoSueldo, id=pago_id)
    
    if request.method == 'POST':
        from decimal import Decimal
        
        # Función blindada
        def a_decimal(valor):
            if not valor or str(valor).strip() == 'None' or str(valor).strip() == '':
                return Decimal('0.00')
            try:
                return Decimal(str(valor).strip())
            except:
                return Decimal('0.00')

        pago.salario_base = a_decimal(request.POST.get('salario_base'))
        pago.bono_antiguedad = a_decimal(request.POST.get('bono_antiguedad'))
        pago.bono_ventas = a_decimal(request.POST.get('bono_ventas'))
        pago.comisiones_certificados = a_decimal(request.POST.get('comisiones_certificados'))
        pago.bono_consultora = a_decimal(request.POST.get('bono_consultora'))
        pago.otros_bonos = a_decimal(request.POST.get('otros_bonos'))
        
        pago.aportes_afp = a_decimal(request.POST.get('aportes_afp'))
        pago.rc_iva = a_decimal(request.POST.get('rc_iva'))
        pago.anticipos = a_decimal(request.POST.get('anticipos'))
        pago.prestamos = a_decimal(request.POST.get('prestamos'))
        pago.multas = a_decimal(request.POST.get('multas'))
        pago.rendicion_cuentas = a_decimal(request.POST.get('rendicion_cuentas'))
        pago.pasanaku = a_decimal(request.POST.get('pasanaku'))
        
        fecha_recibida = request.POST.get('fecha_pago')
        if fecha_recibida and str(fecha_recibida).strip() != '':
            pago.fecha_pago = fecha_recibida
            
        pago.mes_correspondiente = request.POST.get('mes_correspondiente')
        
        cuenta_origen_id = request.POST.get('cuenta_origen')
        if cuenta_origen_id and str(cuenta_origen_id).strip() != '':
            pago.cuenta_origen_id = cuenta_origen_id
        else:
            pago.cuenta_origen_id = None
            
        pago.save()
        messages.success(request, 'El registro de planilla fue actualizado correctamente.')
        return redirect('planillas')
        
    empleados = Empleado.objects.all().order_by('nombre_completo')
    cuentas = CuentaCaja.objects.all().order_by('codigo')
    return render(request, 'crear_pago.html', {
        'pago': pago, 
        'empleados': empleados, 
        'cuentas': cuentas,
        'editando': True
    })
@login_required
@user_passes_test(es_administrador)
def crear_cliente(request):
    if request.method == 'POST':
        accion = request.POST.get('accion')
        cliente_id = request.POST.get('cliente_id')
        
        # 1. ACCIÓN: ELIMINAR (CON CANDADO DE SEGURIDAD PROTECTEDERROR)
        if accion == 'eliminar' and cliente_id:
            try:
                cliente = get_object_or_404(Cliente, id=cliente_id)
                nombre_temporal = cliente.nombre_contribuyente
                cliente.delete()
                messages.success(request, f'El cliente "{nombre_temporal}" fue eliminado correctamente del sistema.')
            except ProtectedError:
                messages.error(request, 'ACCIÓN BLOQUEADA: No se puede eliminar a este cliente porque ya tiene trámites o servicios registrados en su historial. Primero debe anular sus trámites.')
            return redirect('crear_cliente')
            
        # Extraemos los datos del formulario (comunes para registrar y actualizar)
        nit = request.POST.get('nit', '')
        nombre_contribuyente = request.POST.get('nombre_contribuyente', '').upper()
        domicilio_fiscal = request.POST.get('domicilio_fiscal', '')
        celular = request.POST.get('celular', '')
        correo = request.POST.get('correo', '')
        denominacion = request.POST.get('denominacion', '')
        contrasena = request.POST.get('contrasena', '')

        # 2. ACCIÓN: ACTUALIZAR
        if accion == 'actualizar' and cliente_id:
            cliente = get_object_or_404(Cliente, id=cliente_id)
            cliente.nit = nit
            cliente.nombre_contribuyente = nombre_contribuyente
            cliente.domicilio_fiscal = domicilio_fiscal
            cliente.celular = celular
            cliente.correo = correo
            cliente.denominacion = denominacion
            cliente.contrasena = contrasena
            cliente.save()
            messages.success(request, 'Los datos del cliente fueron actualizados con éxito.')
            return redirect('crear_cliente')
            
        # 3. ACCIÓN: REGISTRAR NUEVO
        elif accion == 'registrar':
            Cliente.objects.create(
                nit=nit,
                nombre_contribuyente=nombre_contribuyente,
                domicilio_fiscal=domicilio_fiscal,
                celular=celular,
                correo=correo,
                denominacion=denominacion,
                contrasena=contrasena
            )
            messages.success(request, 'El nuevo cliente fue registrado oficialmente.')
            return redirect('crear_cliente')

    # 4. Cargamos la base de datos completa y la enviamos al HTML para el buscador
    todos_los_clientes = Cliente.objects.all().order_by('nombre_contribuyente')
    
    return render(request, 'crear_cliente.html', {'clientes': todos_los_clientes})
# --- NUEVO: API DE BÚSQUEDA INTELIGENTE DE NITS ---
@login_required
def api_buscar_cliente(request):
    q = request.GET.get('q', '')
    resultados = []
    
    # 1. SI ESCRIBE ALGO, BUSCA COINCIDENCIAS EXACTAS
    if q:
        clientes = Cliente.objects.filter(Q(nit__icontains=q) | Q(nombre_contribuyente__icontains=q))[:20]
    # 2. SI ESTÁ VACÍO (Al hacer clic en el desplegable), TRAE TODOS
    else:
        clientes = Cliente.objects.all().order_by('-id')[:100] # Traemos los últimos 100 para no colapsar la memoria
        
    for c in clientes:
        resultados.append({
            'id': c.id,
            'nit': c.nit or '',
            'nombre_contribuyente': c.nombre_contribuyente or '',
            'domicilio_fiscal': c.domicilio_fiscal or '',
            'celular': c.celular or '',
            'correo': c.correo or '',
            'denominacion': c.denominacion or '',
            'contrasena': c.contrasena or ''
        })
    return JsonResponse(resultados, safe=False)

@login_required
def lista_prestamos(request):
    buscar = request.GET.get('buscar', '').strip()
    
    # ¡OPTIMIZACIÓN EXTREMA!: Agregamos prefetch_related('pagos') para traer el historial en 1 solo viaje
    prestamos = Prestamo.objects.select_related('empleado').prefetch_related('pagos').all().order_by('-fecha_prestamo')
    
    # --- 1. MOTOR DE BÚSQUEDA POR TEXTO (Nombre o C.I.) ---
    if buscar:
        prestamos = prestamos.filter(
            Q(empleado__nombre_completo__icontains=buscar) |
            Q(empleado__ci__icontains=buscar)
        )
        
    # --- 2. CÁLCULO DE TOTALES PARA LAS TARJETAS (KPIs) ---
    totales_bd = prestamos.aggregate(
        t_capital=Sum('monto_prestado'),
        t_pagos=Sum('pagos__monto') # Suma todos los pagos asociados
    )
    
    total_capital_prestado = totales_bd['t_capital'] or Decimal('0.00')
    total_recuperado = totales_bd['t_pagos'] or Decimal('0.00')
    saldo_por_cobrar = total_capital_prestado - total_recuperado

    contexto = {
        'prestamos': prestamos,
        'buscar': buscar,
        'total_capital_prestado': total_capital_prestado,
        'total_recuperado': total_recuperado,
        'saldo_por_cobrar': saldo_por_cobrar,
    }
    return render(request, 'lista_prestamos.html', contexto)

@login_required
@user_passes_test(es_administrador)
def crear_prestamo(request):
    if request.method == 'POST':
        empleado_raw = request.POST.get('empleado')
        # ¡SOLUCIÓN!: Si viene vacío (es externo), lo convertimos a None para que Django no colapse
        empleado_id = int(empleado_raw) if empleado_raw else None
        
        # Atrapamos los datos del externo
        nombre_externo = request.POST.get('nombre_externo', '').upper()
        celular_externo = request.POST.get('celular_externo', '')
        
        # ... (código anterior: nombre_externo, celular_externo, etc.) ...
        
        monto = Decimal(request.POST.get('monto_prestado', '0') or '0')
        
        # --- NUEVA LÓGICA: CÁLCULO POR PORCENTAJE ---
        porcentaje_interes = Decimal(request.POST.get('interes_porcentaje', '0') or '0')
        # Calculamos el dinero real: Monto * (Porcentaje / 100)
        interes_monto = monto * (porcentaje_interes / Decimal('100'))
        
        cuenta_id = request.POST.get('cuenta_origen_id')

        prestamo = Prestamo(
            empleado_id=empleado_id,
            nombre_externo=nombre_externo,
            celular_externo=celular_externo,
            fecha_prestamo=request.POST.get('fecha_prestamo'),
            tipo_prestamo=request.POST.get('tipo_prestamo', 'AMORTIZABLE'),
            tipo_cuota=request.POST.get('tipo_cuota', 'MENSUAL'),
            dia_de_pago=request.POST.get('dia_de_pago', '').upper(),
            monto_prestado=monto,
            interes=interes_monto, # Guardamos el dinero calculado
            total_deuda=monto + interes_monto, # Capital + Interés (El sistema lo dividirá en las cuotas)
            nro_cuotas=request.POST.get('nro_cuotas', '1'),
            observaciones=request.POST.get('observaciones', '')
        )
        prestamo.save()
        
        # ... (resto del código de MovimientoCaja y redirección se mantiene igual) ...# Esto dispara el save() del modelo

        # --- RE-DIRECCIONAMIENTO INTELIGENTE DEL MOVIMIENTO DE CAJA ---
        if cuenta_id:
            try:
                cuenta_real = CuentaCaja.objects.get(id=cuenta_id)
                movimiento = MovimientoCaja.objects.filter(
                    fecha=prestamo.fecha_prestamo,
                    tipo='SALIDA',
                    monto=monto
                ).last()
                if movimiento:
                    movimiento.cuenta = cuenta_real
                    movimiento.save()
            except Exception as e:
                print(f"Error al redireccionar cuenta de préstamo: {e}")

        messages.success(request, f'Préstamo otorgado con éxito. Se registró la salida de Bs. {monto}.')
        return redirect('lista_prestamos')
        
    # GET: Enviamos empleados y cuentas ordenadas por código
    empleados = Empleado.objects.all().order_by('nombre_completo')
    cuentas = CuentaCaja.objects.all().order_by('codigo')
    return render(request, 'crear_prestamo.html', {'empleados': empleados, 'cuentas': cuentas})


@login_required
@user_passes_test(es_administrador)
def editar_prestamo(request, prestamo_id):
    prestamo = get_object_or_404(Prestamo, id=prestamo_id)
    
    if request.method == 'POST':
        # --- FUNCIÓN 1: ELIMINAR REGISTRO ---
        if 'eliminar' in request.POST:
            prestamo.delete()
            messages.success(request, 'El registro del préstamo y su historial de pagos han sido eliminados correctamente.')
            return redirect('lista_prestamos')

        # --- FUNCIÓN 2: ACTUALIZAR CONDICIONES Y REGISTRAR NUEVO DESEMBOLSO ---
        aumento_capital = Decimal(request.POST.get('aumento_capital', '0').strip() or '0')
        aumento_interes = Decimal(request.POST.get('aumento_interes', '0').strip() or '0')
        nueva_fecha = request.POST.get('fecha_prestamo')
        cuenta_id = request.POST.get('cuenta_origen_id')
        fecha_ampliacion = request.POST.get('fecha_ampliacion') # <-- NUEVO: Atrapamos la fecha de la ampliación

        # El préstamo principal se queda con la fecha de reajuste general
        prestamo.fecha_prestamo = nueva_fecha
        prestamo.tipo_cuota = request.POST.get('tipo_cuota')
        prestamo.dia_de_pago = request.POST.get('dia_de_pago', '').upper()
        prestamo.nro_cuotas = request.POST.get('nro_cuotas', 1)
        prestamo.observaciones = request.POST.get('observaciones', '')
        
        if aumento_capital > 0 or aumento_interes > 0:
            prestamo.monto_prestado += aumento_capital
            prestamo.interes += aumento_interes
            prestamo.total_deuda += (aumento_capital + aumento_interes)
            
            # AUTOMATIZACIÓN: Si se inyecta nuevo capital físico, sale de la cuenta en la fecha de la ampliación
            if aumento_capital > 0:
                try:
                    cuenta_origen = CuentaCaja.objects.get(id=cuenta_id)
                except Exception:
                    cuenta_origen, _ = CuentaCaja.objects.get_or_create(codigo='001', defaults={'nombre': 'ADMINISTRACIÓN'})
                
                # ¡SOLUCIÓN!: Usamos fecha_ampliacion si existe, de lo contrario usamos la fecha general como respaldo
                fecha_asiento = fecha_ampliacion if fecha_ampliacion else nueva_fecha
                
                MovimientoCaja.objects.create(
                    fecha=fecha_asiento, # <-- Aplica la fecha del nuevo desembolso al flujo de caja
                    detalle=f"Ampliación Préstamo: {prestamo.empleado.nombre_completo}",
                    cuenta=cuenta_origen,
                    tipo='SALIDA',
                    monto=aumento_capital
                )
            messages.success(request, f'Ampliación registrada con éxito. Se asentó la salida de Bs. {aumento_capital} en el flujo con fecha {fecha_asiento}.')
        else:
            messages.success(request, f'Condiciones del préstamo de {prestamo.empleado.nombre_completo} actualizadas con éxito.')
        
        if prestamo.saldo_restante <= 0:
            prestamo.estado = 'PAGADO'
        else:
            prestamo.estado = 'ACTIVO'
            
        prestamo.save()
        return redirect('lista_prestamos')
        
    empleados = Empleado.objects.all().order_by('nombre_completo')
    cuentas = CuentaCaja.objects.all().order_by('codigo')
    contexto = {
        'prestamo': prestamo,
        'empleados': empleados,
        'cuentas': cuentas
    }
    return render(request, 'editar_prestamo.html', contexto)
@login_required
@user_passes_test(es_administrador)
def registrar_pago_prestamo(request, prestamo_id):
    prestamo = get_object_or_404(Prestamo, id=prestamo_id)
    
    if request.method == 'POST':
        monto_abono = Decimal(request.POST.get('monto', 0))
        cuenta_id = request.POST.get('cuenta_destino_id') # <-- NUEVO: Atrapamos la cuenta elegida
        
        # Validación de seguridad para evitar sobrepagos
        if monto_abono > prestamo.saldo_restante:
            messages.error(request, f'Error: El monto ingresado (Bs. {monto_abono}) supera al saldo restante de la deuda.')
            return redirect('lista_prestamos')
            
        pago = PagoPrestamo(
            prestamo=prestamo,
            fecha_pago=request.POST.get('fecha_pago'),
            monto=monto_abono,
            es_descuento_planilla=False # Es un pago manual en caja
        )
        pago.save() # Esto guarda el abono y dispara el save() base del modelo en la cuenta 001

        # --- RE-DIRECCIONAMIENTO INTELIGENTE DEL MOVIMIENTO DE CAJA ---
        if cuenta_id:
            try:
                cuenta_real = CuentaCaja.objects.get(id=cuenta_id)
                # Buscamos el asiento de ingreso que creó el modelo automáticamente para asignarle la cuenta elegida
                movimiento = MovimientoCaja.objects.filter(
                    fecha=pago.fecha_pago,
                    detalle=f"Cobro Cuota Préstamo: {prestamo.empleado.nombre_completo}",
                    tipo='ENTRADA',
                    monto=monto_abono
                ).last()
                if movimiento:
                    movimiento.cuenta = cuenta_real
                    movimiento.save()
            except Exception as e:
                print(f"Error al redireccionar cuenta de abono: {e}")

        messages.success(request, f'Abono de Bs. {monto_abono} asentado correctamente en la cuenta seleccionada.')
        return redirect('lista_prestamos')
        
    # GET: Enviamos el préstamo y el catálogo de cuentas completo ordenado por código
    cuentas = CuentaCaja.objects.all().order_by('codigo') # <-- NUEVO
    return render(request, 'registrar_pago_prestamo.html', {'prestamo': prestamo, 'cuentas': cuentas})


@login_required
@user_passes_test(es_administrador)
def imprimir_boleta(request, pago_id):
    pago = get_object_or_404(PagoSueldo, id=pago_id)
    empresa = DatosEmpresa.objects.first()

    # Cálculo preciso de la Antigüedad
    antiguedad_str = "Sin registro de ingreso"
    if pago.empleado.fecha_ingreso and pago.fecha_pago:
        dias_totales = (pago.fecha_pago - pago.empleado.fecha_ingreso).days
        if dias_totales >= 0:
            anios = dias_totales // 365
            meses = (dias_totales % 365) // 30
            dias = (dias_totales % 365) % 30
            antiguedad_str = f"{anios} Años, {meses} Meses, {dias} Días"

    # --- NUEVO: Traductor de mes a formato literal ---
    meses_dict = {
        '01': 'ENERO', '02': 'FEBRERO', '03': 'MARZO',
        '04': 'ABRIL', '05': 'MAYO', '06': 'JUNIO',
        '07': 'JULIO', '08': 'AGOSTO', '09': 'SEPTIEMBRE',
        '10': 'OCTUBRE', '11': 'NOVIEMBRE', '12': 'DICIEMBRE'
    }
    try:
        # Separa "2026-06" en "2026" y "06", luego lo traduce
        anio, mes_num = str(pago.mes_correspondiente).strip().split('-')
        mes_literal = f"{meses_dict.get(mes_num, '')} {anio}"
    except:
        # Por si acaso el dato viene en otro formato
        mes_literal = str(pago.mes_correspondiente).upper()

    contexto = {
        'pago': pago,
        'empresa': empresa,
        'antiguedad': antiguedad_str,
        'copias': ['EMPRESA', 'EMPLEADO'],
        'mes_literal': mes_literal, # <--- Se lo enviamos al HTML
    }

    template = get_template('boleta_pago_pdf.html')
    html = template.render(contexto)

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="Boleta_{pago.empleado.nombre_completo}_{pago.mes_correspondiente}.pdf"'

    pisa_status = pisa.CreatePDF(html, dest=response)

    if pisa_status.err:
        return HttpResponse('Error crítico al compilar la boleta PDF.')

    return response

@login_required
@user_passes_test(es_administrador)
def buscar_boletas(request):
    buscar = request.GET.get('buscar', '').strip()
    mes_buscar = request.GET.get('mes_buscar', '').strip() # <-- NUEVO: Atrapa el mes
    
    # Empezamos trayendo todos los pagos
    pagos = PagoSueldo.objects.select_related('empleado').all().order_by('-fecha_pago')
    
    # 1. Si escribió un nombre o CI, filtramos
    if buscar:
        pagos = pagos.filter(
            Q(empleado__nombre_completo__icontains=buscar) |
            Q(empleado__ci__icontains=buscar)
        )
        
    # 2. Si seleccionó un mes (Ej: "2026-06"), filtramos
    if mes_buscar:
        pagos = pagos.filter(mes_correspondiente=mes_buscar)
        
    # 3. Si no usó NINGÚN filtro, mostramos solo los últimos 15 para no saturar la pantalla
    if not buscar and not mes_buscar:
        pagos = pagos[:15]
        
    return render(request, 'buscar_boletas.html', {
        'pagos': pagos, 
        'buscar': buscar,
        'mes_buscar': mes_buscar # <-- Lo pasamos al HTML para que no se borre al buscar
    })

@login_required
@user_passes_test(es_administrador)
def imprimir_recibo_pago(request, pago_id):
    pago = get_object_or_404(PagoPrestamo, id=pago_id)
    empresa = DatosEmpresa.objects.first()
    
    # --- CORRECCIÓN: Definimos la variable prestamo correctamente desde el pago ---
    prestamo = pago.prestamo
    
    # Determinamos de forma dinámica el saldo anterior al abono
    saldo_anterior = prestamo.saldo_restante + pago.monto
    
    contexto = {
        'pago': pago,
        'prestamo': prestamo,
        'empleado': prestamo.empleado,
        'empresa': empresa,
        'saldo_anterior': saldo_anterior,
        'nuevo_saldo': prestamo.saldo_restante,
    }
    
    template = get_template('recibo_pago_pdf.html')
    html = template.render(contexto)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename=\"Recibo_Cuota_{pago.id:05d}.pdf\"'
    
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('Error crítico al compilar el recibo de caja.')
        
    return response

@login_required
@user_passes_test(es_administrador)
def eliminar_cliente(request, cliente_id):
    cliente = get_object_or_404(Cliente, id=cliente_id)
    try:
        nombre_temporal = cliente.nombre_contribuyente
        cliente.delete()
        messages.success(request, f'El cliente "{nombre_temporal}" fue eliminado exitosamente.')
    except ProtectedError:
        messages.error(request, 'ACCIÓN BLOQUEADA: No se puede eliminar a este cliente porque ya tiene trámites o servicios registrados en su historial. Primero debe anular sus trámites.')
        
    return redirect('crear_cliente')
@login_required
def cuentas_por_cobrar(request):
    # Traemos solo las inscripciones que deben dinero
    cuentas = Inscripcion.objects.filter(saldo_pendiente__gt=0).select_related('participante', 'curso').order_by('-fecha_inscripcion')
    
    # Calculamos el KPI global de cuentas por cobrar
    total_deuda = cuentas.aggregate(total=Sum('saldo_pendiente'))['total'] or 0
    
    return render(request, 'cuentas_por_cobrar.html', {
        'cuentas': cuentas,
        'total_deuda': total_deuda
    })

@login_required
@user_passes_test(es_administrador)
def liquidar_saldo_inscripcion(request, id):
    # Importaciones explícitas dentro del entorno de la función para evitar NameErrors
    from datetime import date
    from decimal import Decimal
    from django.db.models import F
    
    inscripcion = get_object_or_404(Inscripcion, id=id)
    
    if request.method == 'POST':
        monto_pago = Decimal(str(request.POST.get('monto_pago', '0').strip() or '0'))
        
        if monto_pago > 0:
            if monto_pago > inscripcion.saldo_pendiente:
                messages.error(request, 'El monto ingresado supera el saldo pendiente del alumno.')
                return redirect('cuentas_por_cobrar')
            
            # Resguardamos cuánto dinero acumulaba en anticipos antes de procesar este abono
            anticipo_previo = inscripcion.importe
            
            # Ejecutamos la amortización en la base de datos
            inscripcion.importe += monto_pago
            nuevo_saldo = inscripcion.saldo_pendiente - monto_pago
            inscripcion.saldo_pendiente = nuevo_saldo
            inscripcion.save()
            
            # Determinamos las cuentas contables de destino
            banco_limpio = str(inscripcion.banco).strip().upper()
            codigo_cuenta = '001' if banco_limpio in ['ADMINISTRACIÓN', 'ADMINISTRACION', 'EFECTIVO', '001 - ADMINISTRACIÓN'] else '002'
            nombre_cuenta = 'ADMINISTRACIÓN' if codigo_cuenta == '001' else 'BANCO'
            cuenta_caja, _ = CuentaCaja.objects.get_or_create(codigo=codigo_cuenta, defaults={'nombre': nombre_cuenta})
            
            # --- CASO A: EL ALUMNO LIQUIDÓ LA TOTALIDAD DE SU DEUDA (SALDO = 0) ---
            if nuevo_saldo == 0:
                
                # 1. Enviamos el valor del anticipo acumulado a los EGRESOS (SALIDA) - Buscador blindado anti-duplicados
                mov_egreso = MovimientoCaja.objects.filter(
                    fecha=date.today(),
                    cuenta=cuenta_caja,
                    tipo='SALIDA',
                    detalle="ANTICIPO DE INSCRIPCIÓN"
                ).first()
                
                if mov_egreso:
                    MovimientoCaja.objects.filter(id=mov_egreso.id).update(monto=F('monto') + anticipo_previo)
                else:
                    MovimientoCaja.objects.create(
                        fecha=date.today(),
                        cuenta=cuenta_caja,
                        tipo='SALIDA',
                        detalle="ANTICIPO DE INSCRIPCIÓN",
                        monto=anticipo_previo
                    )
                
                # 2. Registramos el VALOR TOTAL COMPLETADO en los INGRESOS (ENTRADA) con formato de modalidad limpio
                detalle_normal = inscripcion.modalidad.upper()
                mov_ingreso = MovimientoCaja.objects.filter(
                    fecha=date.today(),
                    cuenta=cuenta_caja,
                    tipo='ENTRADA',
                    detalle=detalle_normal
                ).first()
                
                if mov_ingreso:
                    MovimientoCaja.objects.filter(id=mov_ingreso.id).update(monto=F('monto') + inscripcion.importe)
                else:
                    MovimientoCaja.objects.create(
                        fecha=date.today(),
                        cuenta=cuenta_caja,
                        tipo='ENTRADA',
                        detalle=detalle_normal,
                        monto=inscripcion.importe
                    )
                
                messages.success(request, f'¡Cobro finalizado con éxito! Se revirtieron Bs. {anticipo_previo} de Anticipos a Egresos y se asentó el ingreso total de Bs. {inscripcion.importe} como {detalle_normal}.')
                return redirect('inscripciones')
                
            # --- CASO B: EL ALUMNO REALIZÓ UN ABONO PARCIAL (TODAVÍA DEBE) ---
            else:
                # Buscador blindado anti-duplicados para acumular el pago parcial en los ingresos de anticipos
                mov_parcial = MovimientoCaja.objects.filter(
                    fecha=date.today(),
                    cuenta=cuenta_caja,
                    tipo='ENTRADA',
                    detalle="ANTICIPO DE INSCRIPCIÓN"
                ).first()
                
                if mov_parcial:
                    MovimientoCaja.objects.filter(id=mov_parcial.id).update(monto=F('monto') + monto_pago)
                else:
                    MovimientoCaja.objects.create(
                        fecha=date.today(),
                        cuenta=cuenta_caja,
                        tipo='ENTRADA',
                        detalle="ANTICIPO DE INSCRIPCIÓN",
                        monto=monto_pago
                    )
                
                messages.success(request, f'Se registró un abono parcial de Bs. {monto_pago}. El alumno aún mantiene un saldo de Bs. {nuevo_saldo}.')
                return redirect('cuentas_por_cobrar')
                
    return redirect('inscripciones')
@login_required
def crear_inscripcion_cc(request):
    if request.method == 'POST':
        nombre_completo = request.POST.get('nombre_completo', '').strip().upper()
        celular = request.POST.get('celular', '').strip()
        
        curso_id = request.POST.get('curso')
        fecha = request.POST.get('fecha_inscripcion')
        
        monto_pagado = Decimal(str(request.POST.get('monto_pagado', '0.00').strip() or '0.00'))
        saldo_pendiente = Decimal(str(request.POST.get('saldo_pendiente', '0.00').strip() or '0.00'))
        
        banco = request.POST.get('banco')
        forma_pago = request.POST.get('forma_pago')
        modalidad = request.POST.get('modalidad', 'VIRTUAL').upper()
        vendedor = request.POST.get('vendedor', '').upper()
        registrado_por = request.POST.get('registrado_por', '').upper()
        
        participante, created = Participante.objects.get_or_create(
            nombre_completo=nombre_completo,
            defaults={'celular': celular}
        )
        if not created and celular and participante.celular != celular:
            participante.celular = celular
            participante.save()
        
        curso_seleccionado = get_object_or_404(Curso, id=curso_id)
        subcursos = curso_seleccionado.subcursos.all().order_by('id')
        
        # =========================================================
        # 🤖 CEREBRO DE DISTRIBUCIÓN MATEMÁTICA PARA DEUDAS
        # =========================================================
        if subcursos.exists():
            cantidad = subcursos.count()
            
            pago_base = round(monto_pagado / cantidad, 2)
            deuda_base = round(saldo_pendiente / cantidad, 2)
            
            pago_acumulado = Decimal('0.00')
            deuda_acumulada = Decimal('0.00')
            
            for i, subcurso in enumerate(subcursos):
                if i == cantidad - 1:
                    pago_final = monto_pagado - pago_acumulado
                    deuda_final = saldo_pendiente - deuda_acumulada
                else:
                    pago_final = pago_base
                    deuda_final = deuda_base
                    pago_acumulado += pago_base
                    deuda_acumulada += deuda_base
                    
                if not Inscripcion.objects.filter(participante=participante, curso=subcurso).exists():
                    Inscripcion.objects.create(
                        participante=participante,
                        curso=subcurso, # Ignora el Módulo, guarda el Curso
                        fecha_inscripcion=fecha,
                        importe=pago_final,
                        saldo_pendiente=deuda_final,
                        banco=banco,
                        forma_pago=forma_pago,
                        modalidad=modalidad,
                        vendedor=vendedor,
                        registrado_por=registrado_por
                    )
            messages.success(request, f'Deuda distribuida e inscrita en los {cantidad} cursos del {curso_seleccionado.nombre}.')
            
        else:
            inscripcion_duplicada = Inscripcion.objects.filter(
                participante=participante, curso=curso_seleccionado, modalidad=modalidad
            ).exists()
            
            if inscripcion_duplicada:
                messages.error(request, f'ACCIÓN BLOQUEADA: El alumno ya tiene deuda registrada en el curso "{curso_seleccionado.nombre}".')
                return redirect('cuentas_por_cobrar')
            
            Inscripcion.objects.create(
                participante=participante,
                curso=curso_seleccionado,
                fecha_inscripcion=fecha,
                importe=monto_pagado,
                saldo_pendiente=saldo_pendiente,
                banco=banco,
                forma_pago=forma_pago,
                modalidad=modalidad,
                vendedor=vendedor,
                registrado_por=registrado_por
            )
            messages.success(request, f'Inscripción con deuda registrada oficialmente.')
            
        return redirect('cuentas_por_cobrar')
        
    else:
        hoy = date.today()
        cursos = Curso.objects.select_related('docente').annotate(
            orden_estado=Case(
                When(fecha_inicio__gt=hoy, then=Value(1)),
                When(fecha_inicio__lte=hoy, fecha_finalizacion__gte=hoy, then=Value(2)),
                default=Value(3),
                output_field=IntegerField(),
            )
        ).order_by('orden_estado', '-fecha_inicio') 
    
    return render(request, 'crear_inscripcion_cc.html', {
        'cursos': cursos,
        'fecha_hoy': date.today().strftime('%Y-%m-%d')
    })

@login_required
def marketing(request):
    from datetime import date, timedelta
    
    # Lógica para guardar la imagen subida
    if request.method == 'POST':
        curso_id = request.POST.get('curso_id')
        imagen = request.FILES.get('imagen_publicidad')
        if curso_id and imagen:
            curso_obj = get_object_or_404(Curso, id=curso_id)
            curso_obj.imagen_publicidad = imagen
            curso_obj.save()
            messages.success(request, f'¡Arte gráfico de "{curso_obj.nombre}" subido exitosamente!')
        return redirect('marketing')

    hoy = date.today()
    # Atrapamos la fecha del nuevo filtro por meses
    mes_buscar = request.GET.get('mes', '') 
    
    cursos_bd = Curso.objects.exclude(fecha_inicio__isnull=True).order_by('-fecha_inicio')
    
    # 1. Aplicamos el Filtro de Meses si el usuario buscó algo
    if mes_buscar:
        try:
            anio, mes = mes_buscar.split('-')
            cursos_bd = cursos_bd.filter(fecha_inicio__year=anio, fecha_inicio__month=mes)
        except ValueError:
            pass
    
    campanas = []
    notificaciones_hoy = [] # Lista mágica para atrapar los que se publican HOY
    
    for c in cursos_bd:
        fecha_pub = c.fecha_inicio - timedelta(days=14)
        dias_para_pub = (fecha_pub - hoy).days
        dias_para_inicio = (c.fecha_inicio - hoy).days
        
        # Asignación de Estados
        if dias_para_inicio < 0:
            estado = 'CERRADA'
            color = 'gray'
        elif dias_para_pub <= 0 and dias_para_inicio >= 0:
            estado = 'EN REDES'
            color = 'emerald'
        elif 0 < dias_para_pub <= 7:
            estado = 'PREPARACIÓN'
            color = 'amber'
        else:
            estado = 'PROGRAMADA'
            color = 'blue'
            
        # 2. SISTEMA DE ALERTAS: Si los días para publicar son exactamente CERO (Hoy)
        if dias_para_pub == 0:
            notificaciones_hoy.append(c)
            
        # Asignación de Estados
        if dias_para_inicio < 0:
            estado = 'CERRADA'
            color = 'gray'
        elif dias_para_pub <= 0 and dias_para_inicio >= 0:
            estado = 'EN REDES'
            color = 'emerald'
        elif 0 < dias_para_pub <= 7:
            estado = 'PREPARACIÓN'
            color = 'amber'
        else:
            estado = 'PROGRAMADA'
            color = 'blue'
            
        # 2. SISTEMA DE ALERTAS: Si los días para publicar son exactamente CERO (Hoy)
        if dias_para_pub == 0:
            # Filtro doble seguridad para la alerta flotante
            if not any(n.id == c.id for n in notificaciones_hoy):
                notificaciones_hoy.append(c)
            
        campanas.append({
            'curso': c,
            'fecha_pub': fecha_pub,
            'estado': estado,
            'color': color,
            'dias_para_pub': dias_para_pub,
            'dias_para_inicio': dias_para_inicio
        })
        
    # ========================================================
    # --- NUEVO: ALGORITMO DE ORDENAMIENTO INTELIGENTE ---
    # ========================================================
    def prioridad_marketing(item):
        dias = item['dias_para_pub']
        if dias == 0:
            return (0, 0) # 1º: ¡PUBLICAR HOY! (Máxima prioridad)
        elif dias > 0:
            return (1, dias) # 2º: Próximos a publicar (El que le faltan menos días va más arriba)
        else:
            if item['estado'] != 'CERRADA':
                return (2, abs(dias)) # 3º: Ya publicados/En redes (Los más recientes arriba)
            else:
                return (3, abs(dias)) # 4º: Cerrados (Al fondo de la tabla)

    # Ordenamos aplicando nuestra regla de prioridad
    campanas = sorted(campanas, key=prioridad_marketing)
    # ========================================================
    
    return render(request, 'marketing.html', {
        'campanas': campanas, 
        'notificaciones_hoy': notificaciones_hoy, 
        'hoy': hoy,
        'mes_buscar': mes_buscar 
    })

@login_required
def arqueo_caja(request):
    # Aquí consultarías los saldos reales de tu modelo de Flujo de Caja
    saldo_admin = 1500.50 
    saldo_banco = 12500.00
    saldo_caja_chica = 500.00
    saldo_gerencia = 4200.00
    saldo_ahorro = 8500.00

    if request.method == 'POST':
        # 1. Capturamos lo que el usuario envió en el formulario
        cuenta_seleccionada = request.POST.get('cuenta_arqueo')
        saldo_en_sistema = request.POST.get('saldo_sistema')
        total_contado = request.POST.get('total_fisico_oculto')
        diferencia_final = request.POST.get('diferencia_oculta')
        obs = request.POST.get('observaciones')

        # 2. Lo guardamos en la base de datos
        ArqueoCaja.objects.create(
            usuario=request.user,
            cuenta=cuenta_seleccionada,
            saldo_sistema=saldo_en_sistema,
            total_fisico=total_contado,
            diferencia=diferencia_final,
            observaciones=obs
        )

        # 3. Mostramos el mensaje y recargamos
        messages.success(request, f'¡Arqueo de {cuenta_seleccionada} guardado correctamente!')
        return redirect('arqueo_caja')

    return render(request, 'arqueo.html', {
        'saldo_admin': saldo_admin,
        'saldo_banco': saldo_banco,
        'saldo_caja_chica': saldo_caja_chica,
        'saldo_gerencia': saldo_gerencia,
        'saldo_ahorro': saldo_ahorro,
    })

@login_required
@user_passes_test(es_administrador)
def historial_arqueos(request):
    # Traemos todos los arqueos desde el más reciente
    arqueos = ArqueoCaja.objects.select_related('usuario').all().order_by('-fecha_registro')
    
    # Atrapamos los filtros del buscador
    buscar_cuenta = request.GET.get('cuenta', '')
    rango_fechas = request.GET.get('rango_fechas', '')

    # 1. Filtro por tipo de cuenta
    if buscar_cuenta:
        arqueos = arqueos.filter(cuenta=buscar_cuenta)

    # 2. Filtro por fechas
    if rango_fechas:
        if ' a ' in rango_fechas:
            fecha_inicio, fecha_fin = rango_fechas.split(' a ')
            arqueos = arqueos.filter(fecha_registro__date__range=[fecha_inicio, fecha_fin])
        else:
            arqueos = arqueos.filter(fecha_registro__date=rango_fechas)

    contexto = {
        'arqueos': arqueos,
        'buscar_cuenta': buscar_cuenta,
        'rango_fechas': rango_fechas
    }
    
    return render(request, 'historial_arqueos.html', contexto)

@login_required
@user_passes_test(es_administrador)
def eliminar_arqueo(request, arqueo_id):
    arqueo = get_object_or_404(ArqueoCaja, id=arqueo_id)
    arqueo.delete()
    
    messages.success(request, 'El registro de arqueo/conciliación fue eliminado exitosamente del historial.')
    return redirect('historial_arqueos')

@login_required
def generar_certificado_individual(request, inscripcion_id):
    # Traemos solo a este alumno en específico
    inscrito = get_object_or_404(Inscripcion, id=inscripcion_id)
    curso = inscrito.curso
    
    # =========================================================
    # 🧠 NUEVO: SELECTOR INTELIGENTE DE DISEÑO POR DOCENTE
    # =========================================================
    nombre_docente = curso.docente.nombre.lower()
    
    if 'juan' in nombre_docente or 'juanjo' in nombre_docente:
        archivo_fondo = 'certificado_juanjo.jpg'
    elif 'mariana' in nombre_docente:
        archivo_fondo = 'certificado_mariana.jpg'
    elif 'rodrigo' in nombre_docente:
        archivo_fondo = 'certificado_rodrigo.jpg'
    else:
        # Respaldo de seguridad por si en el futuro agregas un docente sin diseño aún
        archivo_fondo = 'certificado.jpg' 
    # =========================================================

    # 1. Rutas de la imagen (usando el archivo seleccionado)
    # ... (código de selección de imagen que ya tienes) ...

    ruta_actual = os.path.dirname(os.path.abspath(__file__))
    ruta_imagen = os.path.join(ruta_actual, 'static', archivo_fondo).replace('\\', '/')
    
    # 🚀 NUEVO: Ruta absoluta de la fuente Montserrat
    ruta_fuente = os.path.join(ruta_actual, 'static', 'Montserrat-Bold.ttf').replace('\\', '/')
    
    template = get_template('pdf_certificados.html')
    contexto = {
        'curso': curso,
        'inscrito': inscrito, # (OJO: en la otra función esto dice 'inscritos')
        'ruta_imagen': ruta_imagen,
        'ruta_fuente': ruta_fuente, # <--- NUEVO: Enviamos la fuente
    }
    
    html = template.render(contexto)
    # ... (resto del código)
    
    response = HttpResponse(content_type='application/pdf')
    # "inline" hace que se abra en una nueva pestaña súper rápido, en vez de obligar a descargar
    nombre_limpio = inscrito.participante.nombre_completo.replace(" ", "_")
    response['Content-Disposition'] = f'inline; filename="Certificado_{nombre_limpio}.pdf"'
    
    # Crear el PDF
    pisa_status = pisa.CreatePDF(html, dest=response)
    
    if pisa_status.err:
        return HttpResponse('Tuvimos algunos errores <pre>' + html + '</pre>')
    return response

@login_required
def lista_cursos_certificados(request):
    # Traemos los cursos ordenados
    cursos = Curso.objects.select_related('docente').all().order_by('-id')
    buscar = request.GET.get('buscar', '')
    mes_busqueda = request.GET.get('mes', '') 
    
    if buscar:
        cursos = cursos.filter(Q(nombre__icontains=buscar) | Q(docente__nombre__icontains=buscar))
    if mes_busqueda:
        try:
            anio, mes = mes_busqueda.split('-')
            cursos = cursos.filter(fecha_inicio__year=anio, fecha_inicio__month=mes)
        except ValueError:
            pass

    return render(request, 'certificados_cursos.html', {
        'cursos': cursos,
        'buscar': buscar,
        'mes_buscar': mes_busqueda
    })

@login_required
def detalle_curso_certificados(request, curso_id):
    # ==============================================================
    # --- EDICIÓN RÁPIDA DE NOMBRES PARA CERTIFICADOS ---
    # ==============================================================
    if request.method == 'POST' and request.POST.get('accion') == 'corregir_nombre':
        p_id = request.POST.get('participante_id')
        nuevo_nombre = request.POST.get('nuevo_nombre', '').strip().upper()
        if p_id and nuevo_nombre:
            participante = Participante.objects.get(id=p_id)
            participante.nombre_completo = nuevo_nombre
            participante.save()
            messages.success(request, f'El nombre se actualizó correctamente a: {nuevo_nombre}')
        return redirect(request.path_info)
        
    curso = get_object_or_404(Curso, id=curso_id)
    total_virtual = Inscripcion.objects.filter(curso=curso, modalidad='VIRTUAL').count()
    total_presencial = Inscripcion.objects.filter(curso=curso, modalidad='PRESENCIAL').count()
    
    modalidad_param = request.GET.get('modalidad')
    if modalidad_param:
        modalidad_actual = modalidad_param.upper()
    else:
        modalidad_actual = 'VIRTUAL' if total_virtual > 0 else 'PRESENCIAL' if total_presencial > 0 else 'VIRTUAL'
        
    inscritos = Inscripcion.objects.filter(curso=curso, modalidad=modalidad_actual).select_related('participante')
    
    buscar = request.GET.get('buscar', '').strip()
    if buscar:
        inscritos = inscritos.filter(participante__nombre_completo__icontains=buscar)

    return render(request, 'certificados_detalle.html', {
        'curso': curso,
        'inscritos': inscritos,
        'modalidad_actual': modalidad_actual,
        'total_virtual': total_virtual,
        'total_presencial': total_presencial,
        'buscar': buscar,
    })

@login_required
def informacion_cursos(request):
    # 1. Traemos la fecha de hoy para calcular el mes y año actuales
    hoy = date.today()
    mes_actual_str = hoy.strftime('%Y-%m') # Genera "2026-07"
    
    # 2. Atrapamos los parámetros del buscador
    buscar = request.GET.get('buscar', '').strip()
    # Si el usuario no seleccionó un mes, forzamos que sea el mes actual
    mes_buscar = request.GET.get('mes', '') 
    
    # Preparamos la consulta base (excluyendo los que no tienen fecha)
    cursos = Curso.objects.select_related('docente').exclude(fecha_inicio__isnull=True).order_by('-fecha_inicio')

    # 3. LÓGICA DE FILTRADO
    if buscar:
        # Si busca por texto (Nombre o Docente), ignoramos el mes para que pueda encontrarlo en todo el historial
        cursos = cursos.filter(
            Q(nombre__icontains=buscar) | 
            Q(docente__nombre__icontains=buscar)
        )
        # Limpiamos mes_buscar para que el calendario no muestre un mes si buscó por texto
        mes_buscar = '' 
        
    elif mes_buscar:
        # Si el usuario eligió un mes específico en el calendario (Ej: "2026-08")
        try:
            anio, mes = mes_buscar.split('-')
            cursos = cursos.filter(fecha_inicio__year=anio, fecha_inicio__month=mes)
        except ValueError:
            pass
            
    else:
        # COMPORTAMIENTO POR DEFECTO: Mostrar solo los del MES ACTUAL
        cursos = cursos.filter(fecha_inicio__year=hoy.year, fecha_inicio__month=hoy.month)
        # Le pasamos esta variable a la plantilla para que el input del calendario diga "Julio 2026"
        mes_buscar = mes_actual_str

    return render(request, 'informacion_cursos.html', {
        'cursos': cursos,
        'buscar': buscar,
        'mes_buscar': mes_buscar
    })

@login_required
def citas_consultora(request):
    hoy = date.today()
    
    # Dividimos las citas en 3 grupos para el diseño inteligente
    citas_hoy = CitaConsultora.objects.filter(fecha=hoy, estado='PENDIENTE').order_by('hora')
    citas_proximas = CitaConsultora.objects.filter(fecha__gt=hoy, estado='PENDIENTE').order_by('fecha', 'hora')
    historial = CitaConsultora.objects.exclude(estado='PENDIENTE').order_by('-fecha', '-hora')[:30] # Últimas 30 finalizadas

    return render(request, 'citas_consultora.html', {
        'citas_hoy': citas_hoy,
        'citas_proximas': citas_proximas,
        'historial': historial,
        'hoy': hoy
    })

@login_required
def guardar_cita(request):
    if request.method == 'POST':
        CitaConsultora.objects.create(
            nombre_cliente=request.POST.get('nombre_cliente', '').upper(),
            celular=request.POST.get('celular', ''),
            fecha=request.POST.get('fecha'),
            hora=request.POST.get('hora'),
            motivo=request.POST.get('motivo', '').upper(),
            # --- NUEVO CAMPO AÑADIDO ---
            modalidad=request.POST.get('modalidad', 'PRESENCIAL').upper()
        )
        messages.success(request, '¡Cita programada exitosamente!')
    return redirect('citas_consultora')

@login_required
def cambiar_estado_cita(request, cita_id, nuevo_estado):
    cita = get_object_or_404(CitaConsultora, id=cita_id)
    if nuevo_estado in ['PENDIENTE', 'REALIZADA', 'CANCELADA']:
        cita.estado = nuevo_estado
        cita.save()
        messages.success(request, f'El estado de la cita fue actualizado a {nuevo_estado}.')
    return redirect('citas_consultora')

@login_required
def eliminar_cita(request, cita_id):
    cita = get_object_or_404(CitaConsultora, id=cita_id)
    cita.delete()
    messages.success(request, 'La cita fue eliminada del sistema.')
    return redirect('citas_consultora')

@login_required
def archivo_digital(request):
    archivos = ArchivoDigital.objects.all().order_by('-periodo', 'categoria')
    return render(request, 'archivo_digital.html', {'archivos': archivos})

@login_required
def guardar_archivo(request):
    if request.method == 'POST':
        ArchivoDigital.objects.create(
            periodo=request.POST.get('periodo'),
            categoria=request.POST.get('categoria'),
            enlace_drive=request.POST.get('enlace_drive'),
            descripcion=request.POST.get('descripcion', '').upper()
        )
        messages.success(request, '¡Carpeta de Drive vinculada exitosamente!')
    return redirect('archivo_digital')

@login_required
def eliminar_archivo(request, archivo_id):
    archivo = get_object_or_404(ArchivoDigital, id=archivo_id)
    archivo.delete()
    messages.success(request, 'Vínculo eliminado del Archivo Digital.')
    return redirect('archivo_digital')

@login_required
def confirmar_envio_certificados(request, curso_id):
    if request.method == 'POST':
        curso = get_object_or_404(Curso, id=curso_id)
        
        # Solo marcamos como enviado y guardamos la fecha
        curso.certificados_enviados = True
        curso.fecha_envio_certificados = date.today()
        curso.save()
        
        messages.success(request, f'¡Excelente! Confirmaste el envío de los certificados del curso "{curso.nombre}". El bono se calculará automáticamente en la planilla de este mes.')
        
    return redirect('lista_cursos_certificados')

@login_required
def lista_anticipos(request):
    anticipos = AnticipoEmpleado.objects.select_related('empleado', 'cuenta_origen').all().order_by('-fecha')
    
    buscar = request.GET.get('buscar', '').strip()
    mes_buscar = request.GET.get('mes_buscar', '').strip()
    
    if buscar:
        anticipos = anticipos.filter(empleado__nombre_completo__icontains=buscar)
    if mes_buscar:
        anticipos = anticipos.filter(mes_descuento=mes_buscar)
        
    total_entregado = sum(a.monto for a in anticipos if a.estado == 'PENDIENTE')
    
    return render(request, 'lista_anticipos.html', {
        'anticipos': anticipos,
        'buscar': buscar,
        'mes_buscar': mes_buscar,
        'total_entregado': total_entregado
    })

@login_required
@user_passes_test(es_administrador)
def crear_anticipo(request):
    if request.method == 'POST':
        AnticipoEmpleado.objects.create(
            empleado_id=request.POST.get('empleado_id'),
            fecha=request.POST.get('fecha'),
            mes_descuento=request.POST.get('mes_descuento'),
            monto=Decimal(request.POST.get('monto', '0')),
            cuenta_origen_id=request.POST.get('cuenta_origen_id'),
            observaciones=request.POST.get('observaciones', '').upper()
        )
        messages.success(request, 'Anticipo registrado y deducido de la Caja exitosamente.')
        return redirect('lista_anticipos')
        
    empleados = Empleado.objects.all().order_by('nombre_completo')
    cuentas = CuentaCaja.objects.all().order_by('codigo')
    return render(request, 'crear_anticipo.html', {'empleados': empleados, 'cuentas': cuentas})

@login_required
@user_passes_test(es_administrador)
def eliminar_anticipo(request, anticipo_id):
    anticipo = get_object_or_404(AnticipoEmpleado, id=anticipo_id)
    if anticipo.estado == 'DESCONTADO':
        messages.error(request, 'No puedes eliminar un anticipo que ya fue cerrado en una planilla.')
    else:
        anticipo.delete()
        messages.success(request, 'Registro eliminado y el dinero retornó al Flujo de Caja.')
    return redirect('lista_anticipos')
@login_required
@user_passes_test(es_administrador)
def eliminar_afiche_marketing(request, curso_id):
    import os
    curso = get_object_or_404(Curso, id=curso_id)
    
    if curso.imagen_publicidad:
        # 1. Borramos el archivo físico de la carpeta del servidor
        if os.path.isfile(curso.imagen_publicidad.path):
            os.remove(curso.imagen_publicidad.path)
            
        # 2. Vaciamos el campo en la base de datos
        curso.imagen_publicidad = None
        curso.save(update_fields=['imagen_publicidad'])
        
        # 3. Disparamos el mensaje de éxito
        messages.success(request, f'El arte gráfico del curso "{curso.nombre}" fue eliminado para liberar espacio.')
        
    return redirect('marketing')